'''
导入科学数据集的元数据，以及数据实体
'''

import os

from openpyxl import load_workbook
import django
import uuid
from pathlib import Path

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "mysite.settings")
if django.VERSION >= (1, 7):  # 自动判断版本
    django.setup()

def import_data(cat_path):

    from django.contrib.gis.geos import Point
    from django.contrib.gis.geos import GEOSGeometry
    from users.models import myuser
    from place.place_name.models import PlaceName


    tmpl_file = Path(__file__).parent
    # cat_path = pathlib.Path('../place_info.xlsx')

    wb = load_workbook(cat_path)

    ws = wb['Sheet']
    rows = ws.max_row

    for i in range(2, rows + 1):
        location_name = ws.cell(row=i, column=5).value
        historical_name =ws.cell(row=i, column=4).value 
        set_time = 1800
        cancel_time = 2000
        lat = float(ws.cell(row=i, column= 9).value)
        lon = float(ws.cell(row=i, column= 10).value)
        content = ws.cell(row=i, column= 4).value
        print(location_name, lat, lon, type(lat), type(lon), content)
        if lat and lon:
            location = Point(lon,lat)
            point = GEOSGeometry(location)
            print(location)
            print(point)
            try:
                location = Point(lat, lon)
                point = GEOSGeometry(location)
                print(location)
                print(point)
            except:
                pass
            PlaceName.objects.create(region='中国',location_name=location_name, historical_name=historical_name,
                                 lon=lon, lat=lat, content=content,location=point,set_time=set_time,cancel_time=cancel_time,is_en=0)
        else:
            pass
            # print(lat,lon)
    print("Create")


if __name__ == '__main__':
    print('aa')
    inws = Path(__file__).parent / 'the_data'
    print(inws)
    print(inws.exists())
    for xfile in inws.rglob('*.xlsx'):
        print(xfile)
        import_data(xfile)
