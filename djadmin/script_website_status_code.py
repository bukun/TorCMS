import time
import requests
import smtplib
from email.mime.text import MIMEText
import pathlib
from openpyxl import load_workbook
import time


def view_status_code():

    cat_path = pathlib.Path('./website_list.xlsx')

    wb = load_workbook(str(cat_path))

    ws = wb['Sheet']
    rows = ws.max_row

    for i in range(2, rows + 1):
        title = ws.cell(row=i, column=1).value
        url = ws.cell(row=i, column=2).value


        response = requests.get(url)
        status_code = response.status_code

        if status_code == 200:
            print(f'{title} 网站正常')
        else:
            content = f'{title} 网站异常, 状态码为 {status_code}, 网址：{url}'
            print(content)
            # send_email(content)

    time.sleep(60) # 60秒后再次运行





# def send_mail(to_list, sub, content, cc=None):
#     '''
#     Sending email via Python.
#     '''
#     sender = SMTP_CFG['name'] + "<" + SMTP_CFG['user'] + ">"
#     msg = MIMEText(content, _subtype='html', _charset='utf-8')
#     msg['Subject'] = sub
#     msg['From'] = sender
#     msg['To'] = ";".join(to_list)
#     if cc:
#         msg['cc'] = ';'.join(cc)
#     try:
#         # Using SMTP_SSL. The alinyun ECS has masked the 25 port since 9,2016.
#         smtper = smtplib.SMTP_SSL(SMTP_CFG['host'], port=994)
#         smtper.login(SMTP_CFG['user'], SMTP_CFG['pass'])
#         smtper.sendmail(sender, to_list, msg.as_string())
#         smtper.close()
#         return True
#     except Exception as err:
#         print(repr(err))
#         return False

#
# def send_email(content):
#     sender = 'sender@example.com'
#     receiver = 'receiver@example.com'
#     smtp_server = 'smtp.example.com'
#
#     message = MIMEText(content, 'plain', 'utf-8')
#     message['From'] = sender
#     message['To'] = receiver
#     message['Subject'] = '网站监控'
#
#     smtp_obj = smtplib.SMTP(smtp_server)
#     smtp_obj.sendmail(sender, receiver, message.as_string())
#     smtp_obj.quit()
if __name__ == '__main__':
    view_status_code()
