# -*- coding:utf-8 -*-

"""
运行方式：

    DJANGO_SETTINGS_MODULE="mysite.settings" && python3 script_import_datasets.py
"""

import sys

sys.path.append('.')
import requests
import os
import pathlib
from openpyxl import load_workbook


import django
import uuid
from pathlib import Path
from faker import Faker

fk = Faker(locale='zh_CN')
# fake = faker.Faker()


# settings.configure()
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "mysite.settings")
# os.environ.setdefault("DJANGO_SETTINGS_MODULE", "mysite.settings")

from mysite import settings
django.setup()

from data.dataset.models import dataset
from pages.page.models import ThePage


# DJANGO_SETTINGS_MODULE="mysite.settings"
os.environ["DJANGO_SETTINGS_MODULE"] = "mysite.settings"
if django.VERSION >= (1, 7):  # 自动判断版本
    django.setup()


# XLSX_FILE = pathlib.Path('../changchun_project.xlsx')

def search_place(place_name):
    # 接口地址
    url = "https://api.map.baidu.com/geocoding/v3"

    # 此处填写你在控制台-应用管理-创建应用后获取的AK
    ak = "LMIgJaMF56UXqY3xcUpKVSniVXtaPquc"

    params = {
        "address": place_name,
        "output": "json",
        "ak": ak,

    }

    response = requests.get(url=url, params=params)
    if response:
        # {'status': 0, 'result': {'location': {'lng': 125.29719795670451, 'lat': 43.850904132074156}, 'precise': 0, 'confidence': 80, 'comprehension': 99, 'level': '门址'}}

        print(response.json())


def import_project():
    '''
    datasetid = models.CharField(blank=True, null=True, max_length=255, verbose_name='数据ID')
    title = models.CharField(blank=True, null=False, max_length=255, verbose_name="标题")
    title_alternate = models.CharField(blank=True, max_length=255, verbose_name="别名")
    topicategory = models.CharField(blank=True, max_length=255, verbose_name="建议学科分类")
    language = models.CharField(blank=True, max_length=255, verbose_name="语言")
    type = models.CharField(blank=True, max_length=255, verbose_name="数据类型")
    format = models.CharField(blank=True, max_length=255, verbose_name="数据格式")
    links = models.CharField(blank=True, max_length=255, verbose_name="链接")
    time_begin = models.CharField(blank=True, max_length=255, verbose_name="开始时间")
    time_end = models.CharField(blank=True, max_length=255, verbose_name="结束时间")
    creator = models.CharField(blank=True, max_length=255, verbose_name="数据创建者")
    publisher = models.CharField(blank=True, max_length=255, verbose_name="数据发布者")
    contributor = models.CharField(blank=True, max_length=255, verbose_name="数据贡献者")
    organization = models.CharField(blank=True, max_length=255, verbose_name="组织机构")
    operateson = models.CharField(blank=True, max_length=255, verbose_name="元数据创建者")
    cnt_md = MDTextField(verbose_name="内容", null=True, blank=True)

    label = models.ManyToManyField(labels, related_name='dataset', verbose_name='标签', blank=True)
    category = models.ForeignKey(categorys, on_delete=models.CASCADE, blank=True, null=True,
                                 related_name='dataset', verbose_name='分类名称')
    date = models.DateTimeField(verbose_name='创建日期', auto_now_add=True)
    user = models.ForeignKey(User, on_delete=models.CASCADE, blank=True, null=True, related_name='dataset',
                             editable=False,
                             verbose_name='用户名')
    view_count = models.IntegerField(blank=True, null=True, default=0, verbose_name="浏览量", editable=False)
    logo = models.ImageField(upload_to='dataset/imgs/', max_length=255, null=True, blank=True,
                             verbose_name="图片")
    file = models.FileField(upload_to='dataset/files/', max_length=255, null=True, blank=True,
                            verbose_name="文件")

    sites = models.ManyToManyField(Site, blank=True, related_name='dataset', verbose_name='Site')
    extinfo = models.JSONField(null=True, default=dict, verbose_name='Extra data in JSON.', blank=True)
    '''

    # wb = load_workbook(str(XLSX_FILE))
    # ws = wb['Sheet']
    # rows = ws.max_row
    for ii in range(1, 10):
        # location=search_place(xmdz)
        defaults = {
            'datasetid': ii * 100 + fk.random_int(1, 3),
            'title': 'xx_Faker: ' + fk.text(max_nb_chars=20),
            'language': fk.language_name(),
            'contributor': 'faker'
            # 'location': location,
        }
        dataset.objects.update_or_create(**defaults)

def impor_page():


    for ii in range(1, 10):
        # location=search_place(xmdz)
        defaults = {

            'title': 'xx_Faker: ' + fk.text(max_nb_chars=20),
            'slug': fk.aba(),
            'cnt_md': fk.text(max_nb_chars=1000),
            'edit_count': 100,

        }
        ThePage.objects.update_or_create(**defaults)

if __name__ == '__main__':
    import_project()
    impor_page()
