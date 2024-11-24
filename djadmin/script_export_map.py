'''
导入科学数据集的元数据，以及数据实体
'''

import os
import json
import pathlib
from openpyxl import load_workbook
import django
from openpyxl import Workbook
import uuid
from pathlib import Path
from datetime import datetime
import sys

sys.path.append('.')
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "mysite.settings")
if django.VERSION >= (1, 7):  # 自动判断版本
    django.setup()


class DateTimeEncoder(json.JSONEncoder):
    def default(self, obj):
        if isinstance(obj, datetime):
            return obj.isoformat()
        return json.JSONEncoder.default(self, obj)


def export_data():
    from qgis.qgis_map.models import qgismap
    from users.models import myuser
    wb = Workbook()

    ws = wb.active
    wb2 = Workbook()

    ws2 = wb2.active
    wb3 = Workbook()

    ws3 = wb3.active
    recs = qgismap.objects.filter(zhongmengcategory__isnull=False)

    ws.append(
        ["id", "mapid", "title", "en_title", "cnt_md", "en_cnt_md", "lat", "lon", "zoom_current",
         "zoom_min", "zoom_max", "layer_name", "url", "path", "host", "name", "date",
         "label", "zhongmengcategory"])
    all_labels = []
    all_categorys = []
    for rec in recs:
        labels = ''

        for label in rec.label.all():
            labels = labels + label.name + ','
            # all_labels.append({'id':label.id,'name':label.name,'order':label.order})
            if label.name not in all_labels:
                all_labels.append(label.name)
        ws.append(
            [rec.id, rec.mapid, rec.title, rec.en_title, rec.cnt_md, rec.en_cnt_md, rec.lat, rec.lon, rec.zoom_current,
             rec.zoom_min, rec.zoom_max, rec.layer_name, rec.url, rec.path, rec.host, rec.name, str(rec.date),
             labels, str(rec.zhongmengcategory)])
        if rec.zhongmengcategory not in all_categorys:
            all_categorys.append(str(rec.zhongmengcategory))
    for new_cat in all_categorys:
        ws3.append([str(new_cat)])
    for new_label in all_labels:
        ws2.append([str(new_label)])

    wb3.save(f"{Path(__file__).parent}/xx_map_category.xlsx")
    wb2.save(f"{Path(__file__).parent}/xx_map_label.xlsx")
    wb.save(f"{Path(__file__).parent}/xx_map_data.xlsx")


if __name__ == '__main__':
    export_data()
