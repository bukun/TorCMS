# -*- coding:utf-8 -*-

"""
create YAML file  for MapProxy.
"""

import sys

sys.path.append('.')
import requests
import os
import pathlib
from openpyxl import load_workbook
import django
import uuid
from pathlib import Path

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "mysite.settings")
if django.VERSION >= (1, 7):  # 自动判断版本
    django.setup()

XLSX_FILE = pathlib.Path('../changchun_project.xlsx')

def search_place(place_name):

    # 接口地址
    url = "https://api.map.baidu.com/geocoding/v3"

    # 此处填写你在控制台-应用管理-创建应用后获取的AK
    ak = "LMIgJaMF56UXqY3xcUpKVSniVXtaPquc"

    params = {
        "address": place_name,
        "output": "json",
        "ak": ak,

    }

    response = requests.get(url=url, params=params)
    if response:
        #{'status': 0, 'result': {'location': {'lng': 125.29719795670451, 'lat': 43.850904132074156}, 'precise': 0, 'confidence': 80, 'comprehension': 99, 'level': '门址'}}

        print(response.json())
def import_project():
    from changchun_project.models import ChangChunProject
    wb = load_workbook(str(XLSX_FILE))
    ws = wb['Sheet']
    rows = ws.max_row
    for i in range(2, rows + 1):

        djh = ws.cell(row=i, column=1).value

        name = ws.cell(row=i, column=2).value
        company = ws.cell(row=i, column=3).value
        jzd = ws.cell(row=i, column=4).value
        xkz = ws.cell(row=i, column=5).value
        hzd = ws.cell(row=i, column=6).value
        dzd = ws.cell(row=i, column=7).value
        xmdz = ws.cell(row=i, column=8).value
        # location=search_place(xmdz)
        ChangChunProject.objects.update_or_create(cadastre_id=djh, defaults={

            'name': name,
            'company': company,
            'jzd': jzd,
            'xkz': xkz,
            'hzd': hzd,
            'dzd': dzd,
            'xmdz': xmdz,
            # 'location': location,
        })


if __name__ == '__main__':
    import_project()
