'''
导入科学数据集的元数据，以及数据实体
'''

import os
import pathlib
from openpyxl import load_workbook
import django
import uuid
from pathlib import Path

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "mysite.settings")
if django.VERSION >= (1, 7):  # 自动判断版本
    django.setup()


def import_data():
    from place_name.models import PlaceName
    from users.models import myuser
    from django.contrib.gis.geos import Point
    from django.contrib.gis.geos import GEOSGeometry

    tmpl_file = Path(__file__).parent
    cat_path = pathlib.Path('../place_en.xlsx')

    wb = load_workbook(str(cat_path))

    ws = wb['国外地名按照英文首字母排列']
    rows = ws.max_row
    col = 2

    for i in range(2, rows + 1):

        region = ws.cell(row=i, column=col).value
        location_name = ws.cell(row=i, column=col+1).value

        lat = ws.cell(row=i, column=col + 2).value
        lon = ws.cell(row=i, column=col + 3).value
        content = ws.cell(row=i, column=col + 4).value
        set_time = ws.cell(row=i, column=col + 5).value
        cancel_time = ws.cell(row=i, column=col + 6).value
        historical_name=''
        if len(str(set_time))>1:
            set_time=set_time
        else:
            set_time=''
        if len(str(cancel_time))>1:
            cancel_time=cancel_time
        else:
            cancel_time=''

        if lat and lon:
            try:
                location = Point(lat, lon)
                point = GEOSGeometry(location)
                print(location)
                print(point)
            except:
                pass
        if location_name:
            PlaceName.objects.create(region=region,location_name=location_name, historical_name=historical_name,
                                     lon=lon, lat=lat, content=content,location=point,set_time=set_time,cancel_time=cancel_time,is_en=1)

    print("Create")


if __name__ == '__main__':
    import_data()
