'''
导入科学数据集的元数据，以及数据实体
'''

import os
import pathlib
from openpyxl import load_workbook
import django
import uuid
from pathlib import Path

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "mysite.settings")
if django.VERSION >= (1, 7):  # 自动判断版本
    django.setup()


def import_data():
    from place_name.models import PlaceName
    from users.models import myuser
    from django.contrib.gis.geos import Point
    from django.contrib.gis.geos import GEOSGeometry

    tmpl_file = Path(__file__).parent
    cat_path = pathlib.Path('../place_info.xlsx')

    wb = load_workbook(str(cat_path))

    ws = wb['国内地名']
    rows = ws.max_row
    col = 2

    for i in range(2, rows + 1):
        his_name_arr=ws.cell(row=i, column=col + 1).value.split('、')
        for his_name in his_name_arr:
            location_name = ws.cell(row=i, column=col).value
            historical_name = his_name
            set_time = ws.cell(row=i, column=col + 2).value
            cancel_time = ws.cell(row=i, column=col + 3).value
            lat = ws.cell(row=i, column=col + 5).value
            lon = ws.cell(row=i, column=col + 4).value
            content = ws.cell(row=i, column=col + 6).value
            if lat and lon:
                try:
                    location = Point(lat, lon)
                    point = GEOSGeometry(location)
                    print(location)
                    print(point)
                except:
                    pass
            PlaceName.objects.create(region='中国',location_name=location_name, historical_name=historical_name,
                                     lon=lon, lat=lat, content=content,location=point,set_time=set_time,cancel_time=cancel_time,is_en=0)

    print("Create")


if __name__ == '__main__':
    import_data()
