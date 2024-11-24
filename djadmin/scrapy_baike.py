#!/usr/bin/env/ python3
# coding=utf-8

import requests
from bs4 import BeautifulSoup
import re
import codecs
from urllib.parse import urljoin, urlparse
import time
import os
import pathlib
from pathlib import Path
from openpyxl import load_workbook
import django


current_path = os.getcwd()
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "mysite.settings")
if django.VERSION >= (1, 7):  # 自动判断版本
    django.setup()


class UrlManager(object):
    def __init__(self):
        self.new_urls = set()
        self.old_urls = set()

    def add_new_url(self, rootUrl):
        if rootUrl == None:
            return
        if rootUrl not in self.new_urls and rootUrl not in self.old_urls:
            self.new_urls.add(rootUrl)

    def hasUrl(self):
        return len(self.new_urls) > 0

    def getURL(self):
        url = self.new_urls.pop()
        self.old_urls.add(url)
        return url

    def add_new_urls(self, links):
        for link in links:
            self.add_new_url(link)


class CrawlerManage(object):
    def requestURL(self, url):
        # req = requests.Request(url = url,headers={'User-Agent':'Mozilla/5.0(Macintosh;IntelMacOSX10_7_0)AppleWebKit/535.11(KHTML,likeGecko)Chrome/17.0.963.56Safari/535.11'})
        r = requests.get(url, headers={
            'User-Agent': 'Mozilla/5.0(Macintosh;IntelMacOSX10_7_0)AppleWebKit/535.11(KHTML,likeGecko)Chrome/17.0.963.56Safari/535.11'})
        if r.status_code != 200:
            return
        return r.encoding, r.text

    def analyzeLiks(self, page_url, content):
        fullNewLinks = set()
        soup = BeautifulSoup(content, 'html.parser')
        links = soup.find_all('a', href=re.compile(r"/item/\S+"))
        for link in links:
            new_url = link['href']
            new_full_url = urljoin(page_url, new_url)
            fullNewLinks.add(new_full_url)
        return fullNewLinks
    # <a target="_blank" href="/item/%E9%84%B1%E9%98%B3%E8%A1%97%E5%B0%8F%E5%AD%A6">鄱阳街小学</a>


class RunService(object):
    def __init__(self, rootUrl):
        self.urlManager = UrlManager()
        self.urlManager.add_new_url(rootUrl)
        self.crawlerManage = CrawlerManage()


    def run(self, num,keyword):

        n = 0
        con = ''
        while self.urlManager.hasUrl():
            url = self.urlManager.getURL()
            encodeing, content = self.crawlerManage.requestURL(url)
            con = con + content
            with codecs.open(current_path + '/xx_xzqh/' + keyword+ ".html", 'w+', encodeing) as f:
                f.write(content)

            links = self.crawlerManage.analyzeLiks(url, content)
            self.urlManager.add_new_urls(links)
            if n == num:
                break
            n = n + 1

        return con


def get_city():
    from place.xzqh.models import XZQH

    wb = load_workbook(current_path + '/ceshi.xlsx')

    ws = wb['Sheet1']
    rows = ws.max_row

    for i in range(4, rows + 1):
        zoning = str(ws.cell(row=i, column=2).value).strip()
        name = str(ws.cell(row=i, column=3).value).strip()
        content = get_content(name)
        print(content)
        rec = XZQH.objects.filter(zoning=zoning)
        if rec:
            XZQH.objects.filter(zoning=zoning).update(content=str(content))
        else:
            XZQH.objects.create(zoning=zoning, name=name, content=content)


def get_content(keyword):
    # https://baike.baidu.com/item/%E5%88%98%E4%BA%A6%E8%8F%B2/136156
    # rootUrl = f"https://baike.baidu.com/item/滦南县"
    rootUrl = f"https://baike.baidu.com/item/{keyword}?fromModule=lemma_search-box"

    runService = RunService(rootUrl)
    startTIme = time.time()
    print('start server')
    runService.run(1000,keyword)
    endTIme = time.time()
    longtime = endTIme - startTIme
    print('server exit(0),use time %s s' % str(longtime))


if __name__ == '__main__':
    get_city()
