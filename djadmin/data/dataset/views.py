import json
import os
import openpyxl
from .models import dataset
from rest_framework import generics
from rest_framework import permissions
from .serializers import DataSerializer
from .permissions import IsOwnerOrReadOnly
from django.contrib.auth import get_user_model
from django_filters import rest_framework
from django.views.generic import DetailView
from data.categorys.models import categorys
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.sites.models import Site
from django.http import HttpResponse
from base.models import get_template
from django.conf import settings

from django.template.context_processors import csrf

parent_template = get_template()
current_site = Site.objects.get_current()
User = get_user_model()


class DataList(generics.ListCreateAPIView):
    queryset = dataset.objects.filter(sites__id=current_site.id)
    serializer_class = DataSerializer
    permission_classes = (permissions.IsAuthenticatedOrReadOnly,)

    filter_backends = (rest_framework.DjangoFilterBackend,)
    filterset_fields = ['title', 'category']

    # 将request.user与author绑定
    def perform_create(self, serializer):
        serializer.save(user=self.request.user)


class DataDetail(generics.RetrieveUpdateDestroyAPIView):
    queryset = dataset.objects.filter(sites__id=current_site.id)
    serializer_class = DataSerializer
    permission_classes = (permissions.IsAuthenticatedOrReadOnly, IsOwnerOrReadOnly)


def DataDetailView(request, dataid):
    # 从url里获取单个任务的pk值，然后查询数据库获得单个对象
    current_site = Site.objects.get_current()
    data = get_object_or_404(dataset, pk=dataid)
    data_cat = categorys.objects.filter(sites__id=current_site.id)

    return render(request, "dataset/data_detail.html",
                  {"data": data, "Category": data_cat, 'site': current_site, 'parent_template': parent_template})


def export_to_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active

    data_id = request.POST.get('dataset_id')

    obj = dataset.objects.filter(id=data_id).first()
    ################按行保存###################
    # 添加表头
    # ws.append(
    #     ['id', '数据ID', '标题', '别名', '建议学科分类', '语言', '数据类型', '数据格式', '链接', '开始时间', '结束时间',
    #      '数据创建者',
    #      '数据发布者', '数据贡献者', '组织机构', '元数据创建者', '内容', '标签', '分类名称', '创建日期', '用户名',
    #      '浏览量', '图片',
    #      '文件', '站点', 'extinfo'
    #      ])
    # labels = ''
    #
    # for label in obj.label.all():
    #     labels = labels + label.name + ','

    # ws.append(
    #     [obj.id, obj.datasetid, obj.title, obj.title_alternate, obj.topicategory, obj.language, obj.type, obj.format,
    #      obj.links, str(obj.time_begin), str(obj.time_end), obj.creator, obj.publisher, obj.contributor,
    #      obj.organization,
    #      obj.operateson, obj.cnt_md, labels, obj.category.name, str(obj.date), obj.user.username, obj.view_count,
    #      obj.logo.path,
    #      obj.file.path,
    #      obj.sites.name, str(obj.extinfo)
    #      ])
    ################按行保存###################

    # 以下为按列保存代码
    labels = ''

    for label in obj.label.all():
        labels = labels + label.name + ','
    name_arr = ['id', '数据ID', '标题', '别名', '建议学科分类', '语言', '数据类型', '数据格式', '链接', '开始时间',
                '结束时间',
                '数据创建者',
                '数据发布者', '数据贡献者', '组织机构', '元数据创建者', '内容', '标签', '分类名称', '创建日期',
                '用户名',
                '浏览量', '图片',
                '文件', '站点', 'extinfo'
                ]
    value_arr = [obj.id, obj.datasetid, obj.title, obj.title_alternate, obj.topicategory, obj.language, obj.type,
                 obj.format,
                 obj.links, str(obj.time_begin), str(obj.time_end), obj.creator, obj.publisher, obj.contributor,
                 obj.organization,
                 obj.operateson, obj.cnt_md, labels, obj.category.name, str(obj.date), obj.user.username,
                 obj.view_count,
                 obj.logo.name if obj.logo else None,
                 obj.file.name if obj.file else None,
                 obj.sites.name, str(obj.extinfo)
                 ]
    ii = 1
    for t_name in name_arr:
        ws.cell(row=ii, column=1).value = t_name
        ws.cell(row=ii, column=2).value = value_arr[ii - 1]

        ii = ii + 1

    save_dir = settings.MEDIA_ROOT + "/dataset/download/"
    if os.path.exists(save_dir):
        pass
    else:
        os.mkdir(save_dir)

    save_path = os.path.join(save_dir, f"{obj.id}_dataset.xlsx")
    wb.save(save_path)
    visit_path = f'/media/dataset/download/{obj.id}_dataset.xlsx'
    return HttpResponse(json.dumps(visit_path), content_type='application/json; charset=utf-8')
