# -*- coding: utf-8
'''
Generate the dic of Python from xlsx file.
'''

import os
import sys
from openpyxl.reader.excel import load_workbook
from .base_crud import crud_path

xlsx_file = './database/meta/info_tags.xlsx'

if os.path.exists(xlsx_file):
    wb = load_workbook(filename=xlsx_file)
else:
    sys.exit(0)

# sheet_arr = ['Sheet1', 'Sheet2', 'Sheet3', 'Sheet4', 'Sheet5']


def build_dir():
    tag_arr = ['add', 'edit', 'view', 'list', 'infolist']
    path_arr = [os.path.join(crud_path, x) for x in tag_arr]
    for wpath in path_arr:
        if os.path.exists(wpath):
            continue
        os.makedirs(wpath)


class_arr = ['E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N',
             'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z']

# class_arr = [chr(x) for x in range(69, 91)]

# Save the sig_name( e_name) in array of each sheet. to keep the order.
# sig_name_arr = {}
# for sheet_name in sheet_arr:



def gen_html_dic():
    '''
    生成 Filter .
    :return:
    '''
    if wb:
        pass
    else:
        return False

    fo = open('xxtmp_html_dic.py', 'w')

    for wk_sheet in wb:
        for jj in class_arr:

            row1_val = wk_sheet['{0}1'.format(jj)].value
            row2_val = wk_sheet['{0}2'.format(jj)].value

            if row1_val and row1_val != '':

                c_name, slug_name = row1_val.split(',')

                tags1 = row2_val.split(',')
                tags1 = [x.strip() for x in tags1]
                tags_dic = {}

                if len(tags1) == 1:
                    tags_dic[1] = row2_val
                    ctr_type = 'text'
                else:
                    for ii in range(len(tags1)):
                        tags_dic[ii + 1] = tags1[ii].strip()
                    ctr_type = 'select'

                fo.write('''html_{0} = {{
                    'en': 'tag_{0}',
                    'zh': '{1}',
                    'dic': {2},
                    'type': '{3}',
                    }}\n'''.format(slug_name, c_name, tags_dic, ctr_type))
    fo.close()

def __get_slug(ws, clumn):
    row1_val = ws['{0}1'.format(clumn)].value
    c_name, slug_name = row1_val.split(',')
    return slug_name


def gen_array_crud():
    if wb:
        pass
    else:
        return False

    papa_id = 0

    with open('xxtmp_array_add_edit_view.py', 'w') as fo_edit:

        for work_sheet in wb:
            kind_sig = str(work_sheet['A1'].value).strip()
            # 逐行遍历
            for row_num in range(3, 1000):
                # 父类
                A_cell_value = work_sheet['A{0}'.format(row_num)].value
                # 子类
                B_cell_val = work_sheet['B{0}'.format(row_num)].value

                if A_cell_value or B_cell_val:
                    pass
                else:
                    break
                if A_cell_value and A_cell_value != '':
                    papa_id = A_cell_value.strip()[1:]

                    u_dic = []

                    for ii in class_arr:
                        cell_val = work_sheet['{0}{1}'.format(ii, row_num)].value
                        if cell_val == 1:
                            u_dic.append('{0}'.format(
                                __get_slug(work_sheet, ii)
                            ))

                    fo_edit.write('dic_{0}00 = {1}\n'.format(papa_id, u_dic))
                    fo_edit.write('kind_{0}00 = "{1}"\n'.format(papa_id, kind_sig))

                if B_cell_val and B_cell_val != '':

                    sun_id = B_cell_val.strip()[1:]

                    app_uid = '{0}{1}'.format(papa_id, sun_id)

                    u_dic = []

                    for ii in class_arr:
                        cell_val = work_sheet['{0}{1}'.format(ii, row_num)].value
                        if cell_val == 1:
                            u_dic.append('{0}'.format(
                                __get_slug(work_sheet, ii)
                            ))

                    fo_edit.write('dic_{0} = {1}\n'.format(app_uid, u_dic))
                    fo_edit.write('kind_{0} = "{1}"\n'.format(app_uid, kind_sig))

# if __name__ == '__main__':
#     gen_html_dic()
#     gen_array_crud()
