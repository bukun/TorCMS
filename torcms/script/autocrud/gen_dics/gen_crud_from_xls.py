# -*- coding: utf-8
'''

'''
import os
import sys
# from torcms.script.autocrud.base_crud import FILTER_COLUMNS
# from torcms.script.autocrud.gen_dics_from_xls import wb

from openpyxl.reader.excel import load_workbook

from torcms.script.autocrud.base_crud import xlsx_file, FILTER_COLUMNS

if os.path.exists(xlsx_file):
    wb = load_workbook(filename=xlsx_file)
else:
    sys.exit(0)


def __get_slug(ws, clumn):
    row1_val = ws['{0}1'.format(clumn)].value
    c_name, slug_name = row1_val.split(',')
    return slug_name


def gen_array_crud():
    if wb:
        pass
    else:
        return False

    papa_id = 0

    with open('xxtmp_array_add_edit_view.py', 'w') as fo_edit:

        for work_sheet in wb:
            kind_sig = str(work_sheet['A1'].value).strip()
            # 逐行遍历
            for row_num in range(3, 1000):
                # 父类
                A_cell_value = work_sheet['A{0}'.format(row_num)].value
                # 子类
                B_cell_val = work_sheet['B{0}'.format(row_num)].value

                if A_cell_value or B_cell_val:
                    pass
                else:
                    break
                if A_cell_value and A_cell_value != '':
                    papa_id = A_cell_value.strip()[1:]

                    u_dic = []

                    for ii in FILTER_COLUMNS:
                        cell_val = work_sheet['{0}{1}'.format(ii, row_num)].value
                        if cell_val == 1:
                            u_dic.append('{0}'.format(
                                __get_slug(work_sheet, ii)
                            ))

                    fo_edit.write('dic_{0}00 = {1}\n'.format(papa_id, u_dic))
                    fo_edit.write('kind_{0}00 = "{1}"\n'.format(papa_id, kind_sig))

                if B_cell_val and B_cell_val != '':

                    sun_id = B_cell_val.strip()[1:]

                    app_uid = '{0}{1}'.format(papa_id, sun_id)

                    u_dic = []

                    for ii in FILTER_COLUMNS:
                        cell_val = work_sheet['{0}{1}'.format(ii, row_num)].value
                        if cell_val == 1:
                            u_dic.append('{0}'.format(
                                __get_slug(work_sheet, ii)
                            ))

                    fo_edit.write('dic_{0} = {1}\n'.format(app_uid, u_dic))
                    fo_edit.write('kind_{0} = "{1}"\n'.format(app_uid, kind_sig))
