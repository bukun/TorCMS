import os
import shutil
from pathlib import Path

import openpyxl

from torcms.model.category_model import MCategory
from torcms.model.post_model import MPost


def clean_name(instr):
    return ''.join(instr.split()).replace('/', '')


def test_dump2xlsx():
    outbase = Path(__file__).parent / 'xx_out'
    pp = MCategory.query_pcat()
    for t in pp:
        print(t.uid, t.name, t.slug, t.pid)

        ss = MCategory.query_sub_cat(t.uid)

        out_cat = os.path.join(outbase, t.kind)
        if os.path.exists(out_cat):
            pass
        else:
            os.makedirs(out_cat)

        out_file = os.path.join(out_cat, t.uid + '-' + clean_name(t.name) + '.xlsx')

        wb = openpyxl.Workbook()

        for s in ss:
            ws = wb.create_sheet(s.uid + '-' + clean_name(s.name))

            print(' ' * 4, s.uid, s.name, s.slug, s.pid, s.kind)

            posts = MPost.query_by_tag(s.uid, s.kind)
            idx = 1
            for post in posts:
                ws.cell(row=idx, column=1).value = post.uid
                ws.cell(row=idx, column=2).value = post.title
                idx = idx + 1

        # 对于空的数据库不会产生另外worksheet.
        # ws = wb['Sheet']
        # wb.remove(ws)
        wb.save(out_file)

    if os.path.exists(outbase):
        shutil.rmtree(outbase)
