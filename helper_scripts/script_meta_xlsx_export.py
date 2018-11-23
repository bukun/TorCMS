# import tornado.web
from torcms.handlers.entity_handler import EntityHandler

from openpyxl import Workbook

import os

from torcms.model.core_tab import TabPost

from torcms.model.core_tab import TabTag
from torcms.model.abc_model import Mabc



class Ext_tag(Mabc):
    '''
    Model for Posts.
    '''

    @staticmethod
    def get_taglist(kind):
        taglist = TabTag.select().where(TabTag.kind == kind)
        return taglist



class Ext_Post(Mabc):
    @staticmethod
    def query_all_bytag(tag, kind='2'):
        recs = TabPost.select().where(
            (TabPost.kind == kind) &
            (TabPost.extinfo['def_cat_uid'] == tag)
        )
        return recs

    @staticmethod
    def query_all_uid(uid):
        recs = TabPost.select().where(
            (TabPost.uid == uid)
        )
        return recs










def export_data(kind):

    filename = 'ResForm'

    out_docx = './xx_out.xlsx'
    if os.path.exists(out_docx):
        try:
            os.remove(out_docx)
        except:
            pass

    kind = kind
    taglist = Ext_tag.get_taglist(kind)
    wb = Workbook()

    for tag in taglist:
        if not str(tag).endswith('00'):
            sig = str(tag)
            postinfos = Ext_Post.query_all_bytag(sig,kind = kind)
            wb.create_sheet(str(tag))
            print(len(postinfos))
            ws = wb[sig]
            row_inx = 0
            col_inx = 1
            if len(postinfos):
                for x in postinfos:
                    ws.cell(row=row_inx + 1, column=col_inx).value = x.uid
                    ws.cell(row=row_inx + 1,column = col_inx + 1).value = x.title
                    row_inx = row_inx + 1
    wb.save(out_docx)
    return filename

if __name__ == '__main__':
    export_data('9')
