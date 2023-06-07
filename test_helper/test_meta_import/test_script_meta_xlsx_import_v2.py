'''
Import metadata of pycsw schema.
'''

import os
import pathlib

from openpyxl import load_workbook
from pathlib import Path
from torcms.handlers.post_handler import update_category
from torcms.model.category_model import MCategory
from torcms.model.post_model import MPost


def chuli_meta(sig, metafile):
    print(metafile)
    wb = load_workbook(str(metafile))
    sheet = wb[wb.sheetnames[0]]
    meta_dic = {}
    for row in sheet.iter_rows():
        the_key = row[0].value
        the_val = row[1].value if row[1].value else ''
        meta_dic[the_key] = the_val
        # print()
    meta_dic['identifier'] = sig
    print(meta_dic)
    return meta_dic


def get_meta(catid, sig):
    '''
    Get metadata of dataset via ID.
    '''
    meta_base = './static/dataset_list'
    if os.path.exists(meta_base):
        pass
    else:
        return False

    pp_data = {'logo': '', 'kind': '9'}
    for wroot, wdirs, wfiles in os.walk(meta_base):
        for wdir in wdirs:
            if wdir.lower().endswith(sig):
                #  Got the dataset of certain ID.
                ds_base = pathlib.Path(os.path.join(wroot, wdir))

                for uu in ds_base.iterdir():
                    if uu.name.endswith('.xlsx'):
                        meta_dic = chuli_meta('u' + sig[2:], uu)
                        pp_data['title'] = meta_dic['title']
                        pp_data['cnt_md'] = meta_dic['anytext']
                        pp_data['user_name'] = 'admin'
                        pp_data['def_cat_uid'] = catid
                        pp_data['gcat0'] = catid
                        pp_data['def_cat_pid'] = catid[:2] + '00'
                        pp_data['extinfo'] = {}
                    elif uu.name.startswith('thumbnail_'):
                        pp_data['logo'] = os.path.join(wroot, wdir, uu.name).strip('.')

    return pp_data


def test_import_meta():
    inws = Path(__file__).parent / 'pycsw_meta'
    out_file = Path(__file__).parent / 'xx_pycsw_meta.xlsx'

    for cat_path in inws.iterdir():
        print('x' * 40)
        print(cat_path.name)
        if cat_path.name.lower().endswith('.xlsx'):
            print('////')
            pass
        else:
            continue
        wb = load_workbook(str(cat_path))
        sheets = wb.sheetnames
        for sheet in sheets:
            catid = sheet.split('_')[0]
            catinfo = MCategory.get_by_uid(catid)
            if catinfo:
                pass
            else:
                continue
            ws = wb[sheet]
            rows = ws.max_row
            for i in range(1, rows):
                sig = ws.cell(row=i, column=1).value
                if sig:
                    print(sig)
                    pp_data = get_meta(catid, sig)

                    MPost.add_or_update('u' + sig[2:], pp_data)
                    update_category('u' + sig[2:], pp_data)

                    ws.cell(row=i, column=2).value = pp_data['title']
        # Write a copy with titles.
        wb.save(out_file)
    os.remove(out_file)


if __name__ == '__main__':
    test_import_meta()
