'''
导入科学数据集的元数据，以及数据实体
'''

import os
import pathlib
from pathlib import Path
from openpyxl import load_workbook
from torcms.model.category_model import MCategory
from torcms.model.post_model import MPost
from torcms.core.tools import get_uu8d
from torcms.handlers.post_handler import update_category
from torcms.model.post2catalog_model import MPost2Catalog
from torcms.model.label_model import MPost2Label
import shutil
from torcms.model.core_tab import TabPost
from torcms.core import tools
import time
from datetime import datetime
import tornado.escape


def chuli_meta(sig, metafile):
    wb = load_workbook(str(metafile))
    sheet = wb[wb.sheetnames[0]]
    meta_dic = {}

    ii = 1
    for row in sheet.iter_rows():
        ii = ii + 1
        if row[0].value in ['date', 'date_revision', 'date_creation', 'date_publication', 'date_modified',
                            'insert_date', 'time_begin', 'time_end']:
            meta_dic['pycsw_' + str(row[0].value).strip()] = str(row[1].value)[:10]
        else:
            meta_dic['pycsw_' + str(row[0].value).strip()] = str(row[1].value)
        meta_dic['pycsw_identifier'] = sig
    return meta_dic


def get_meta(catid, sig, dataset_id):
    '''
    Get metadata of dataset via ID.
    '''
    meta_base = './static/dataset_list'
    if os.path.exists(meta_base):
        pass
    else:
        # print("The datasets path does not exists.")
        return False
    pp_data = {'logo': '', 'kind': '9'}
    for wroot, wdirs, wfiles in os.walk(meta_base):
        # print(wdirs)
        for wdir in wdirs:
            if wdir.lower().endswith(dataset_id):
                #  Got the dataset of certain ID.
                ds_base = pathlib.Path(os.path.join(wroot, wdir))
                new_extinfo = {}
                for uu in ds_base.iterdir():
                    qian, hou = os.path.splitext(uu.name)
                    if uu.name.startswith('meta') and uu.name.endswith('.xlsx'):
                        # print(uu.name)
                        meta_dics = chuli_meta('9' + dataset_id[2:], uu)

                        if 'pycsw_title' in meta_dics:
                            pass
                        else:
                            continue
                        pp_data['title'] = meta_dics['pycsw_title']
                        pp_data['cnt_md'] = meta_dics['pycsw_abstract']
                        pp_data['user_name'] = 'admin'
                        pp_data['tags'] = meta_dics['pycsw_keywords'] + ',_turkey'
                        pp_data['def_cat_uid'] = catid
                        pp_data['gcat0'] = catid
                        pp_data['def_cat_pid'] = catid[:2] + '00'
                        pp_data['tag_data_format'] = ''
                        pp_data['tag_subject'] = ''
                        for meta_dic in meta_dics:
                            new_extinfo[meta_dic] = meta_dics[meta_dic]



                    elif uu.name.startswith('thumbnail_') and hou.lower() in ['.jpg', '.png', '.gif']:
                        pp_data['logo'] = os.path.join(wroot, wdir, uu.name).strip('.')

                postinfo = MPost.get_by_uid(sig)

                if postinfo:
                    postinfo.extinfo.update(new_extinfo)
                    pp_data['extinfo'] = postinfo.extinfo
                else:
                    pp_data['extinfo'] = new_extinfo
                # print(pp_data)

                # update_post(sig,pp_data,new_extinfo)
                # update_category(sig, pp_data)
                # update_label(sig, pp_data)
                # MPost2Catalog.add_record(sig, catid)

    return pp_data


def update_post(uid, post_data, cur_extinfo):
    '''
    注意，不更新 title
    '''
    entry = TabPost.update(
        # title=post_data['title'].strip(),
        # user_name=data_dic['user_name'],
        keywords=post_data['keywords'] if 'keywords' in post_data else '',
        time_update=tools.timestamp(),
        date=datetime.now(),
        # cnt_md=post_data['cnt_md'],
        # cnt_md=tornado.escape.xhtml_escape(post_data['cnt_md'].strip()),
        memo=post_data.get('memo', ''),
        logo=post_data.get('logo', '').strip(),
        order=post_data.get('order', ''),
        cnt_html=tools.markdown2html(post_data['cnt_md']),
        extinfo=cur_extinfo,
        valid=post_data.get('valid', 1)
    ).where(TabPost.uid == uid)
    entry.execute()


def update_label(signature, post_data):
    '''
    Update the label .
    '''
    current_tag_infos = MPost2Label.get_by_uid(signature).objects()
    if 'tags' in post_data:
        pass
    else:
        return False

    if '；' in post_data['tags']:
        tags_arr = [x.strip() for x in post_data['tags'].split('；')]
    elif ',' in post_data['tags']:
        tags_arr = [x.strip() for x in post_data['tags'].split(',')]
    elif '，' in post_data['tags']:
        tags_arr = [x.strip() for x in post_data['tags'].split('，')]
    elif ';' in post_data['tags']:
        tags_arr = [x.strip() for x in post_data['tags'].split(';')]
    else:
        tags_arr = [x.strip() for x in post_data['tags'].split(',')]

    if len(tags_arr) > 5:
        del tags_arr[5:]

    for tag_name in tags_arr:
        if tag_name == '':
            pass
        else:
            MPost2Label.add_record(signature, tag_name, 1)

    for cur_info in current_tag_infos:
        if cur_info.tag_name in tags_arr:
            pass
        else:
            MPost2Label.remove_relation(signature, cur_info.tag_id)


def check_ds(dataid):
    '''
    检查数据集是否已有
    '''
    inws = Path(__file__).parent / 'datasets'
    for wfile in os.listdir(inws):
        # print(wfile)
        if dataid in wfile:
            return os.path.join(inws, wfile).strip('.')
    return False


def fix_entity_path(data_id):
    '''
    传入数据的 ID 。
    在指定的文件夹下面遍历查找 。 找到后返回路径。
    '''
    # 新拷贝到的路径
    out_dir = Path(__file__).parent / 'datasets'

    if os.path.exists(out_dir):
        pass
    else:
        os.makedirs(out_dir)

    checkit = check_ds(data_id)
    if checkit:
        return checkit
    # 原始数据集路径
    meta_base = '/pb1/ws'

    for wfile in pathlib.Path(meta_base).rglob('entity*.7z'):
        test_sig = wfile.stem.split('_')[-1]
        # print(test_sig)
        if data_id == test_sig:
            out_entity = os.path.join(out_dir, 'entityx_' + get_uu8d() + '_' + data_id + '.7z')

            outpath = out_entity.strip('.')
            # print(f'Out path: {outpath}')
            print(f'Copy {wfile} to {out_entity}')
            # shutil.copy(wfile, out_entity)
            os.symlink(
                wfile,
                os.path.join(os.getcwd(), out_entity)
            )
            return outpath

    return False


def check_post_entity(dataset_id):
    cur_post = MPost.get_by_uid('9' + dataset_id[2:])
    if cur_post:
        entity_path = cur_post.extinfo.get('tag__file_download')
        print('555555555555555555555555')
        print(entity_path)
    else:
        entity_path = ''
    if entity_path:
        entity_path = '.' + entity_path
        if os.path.exists(entity_path):
            print('youa')
            print(entity_path)
            return True
    else:
        print('meiyou')
        pass

    return False


def test_import_meta():
    cat_path = Path(__file__).parent / 'pycsw_meta/pycsw_meta_new.xlsx'
    # print(inws)

    wb = load_workbook(str(cat_path))
    sheets = wb.sheetnames
    for sheet in sheets:
        catid = sheet.split('-')[0].strip()
        catinfo = MCategory.get_by_uid(catid)
        if catinfo:
            pass
        else:
            continue
        ws = wb[sheet]
        rows = ws.max_row
        for i in range(1, rows + 1):
            print('=' * 40)
            uid = ws.cell(row=i, column=2).value
            print(uid)

            if uid:
                pass
            else:
                continue

            dataset_id = ws.cell(row=i, column=1).value

            # 测试 ID
            if dataset_id and dataset_id != '':
                dataset_id = dataset_id
                pass
            else:
                continue

            if uid:
                print("-" * 50)
                print(dataset_id)
                pp_data = get_meta(catid, str(uid), dataset_id)

            if check_post_entity(dataset_id):
                continue

            the_entity_path = fix_entity_path(dataset_id)

            if the_entity_path:
                print(pp_data.keys())
                pp_data['extinfo']['tag__file_download'] = the_entity_path
                # print('Got it')
                MPost.update_jsonb(uid, pp_data['extinfo'])
            else:
                # 如果未找到实体
                print(f'Not found entity: {uid}')
