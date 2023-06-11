# -*- coding: utf-8

'''
导入数据集的信息

 import_meta()从dataset_list调用信息生成数据集
 import_tables()从datacn_tables调用信息生成数据表
'''
import os
import pathlib
import re
import shutil
import sys

from openpyxl import load_workbook
from PIL import Image

from torcms.core.tools import get_uu8d
from torcms.model.category_model import MCategory
from torcms.model.label_model import MPost2Label
from torcms.model.post2catalog_model import MPost2Catalog
from torcms.model.post_model import MPost

logo_cache_dir = './static/cache'


def update_category(uid, postdata, kwargs):
    '''
    Update the category of the post.
    '''
    catid = kwargs['catid'] if ('catid' in kwargs and MCategory.get_by_uid(kwargs['catid'])) else None

    post_data = postdata

    current_infos = MPost2Catalog.query_by_entity_uid(uid, kind='').objects()

    new_category_arr = []
    # Used to update post2category, to keep order.
    def_cate_arr = ['gcat{0}'.format(x) for x in range(10)]

    # for old page.
    def_cate_arr.append('def_cat_uid')

    # Used to update post extinfo.
    cat_dic = {}
    for key in def_cate_arr:
        if key not in post_data:
            continue
        if post_data[key] == '' or post_data[key] == '0':
            continue
        # 有可能选重复了。保留前面的
        if post_data[key] in new_category_arr:
            continue

        new_category_arr.append(post_data[key] + ' ' * (4 - len(post_data[key])))
        cat_dic[key] = post_data[key] + ' ' * (4 - len(post_data[key]))

    if catid:
        def_cat_id = catid
    elif new_category_arr:
        def_cat_id = new_category_arr[0]
    else:
        def_cat_id = None

    if def_cat_id:
        cat_dic['def_cat_uid'] = def_cat_id
        cat_dic['def_cat_pid'] = MCategory.get_by_uid(def_cat_id).pid

    MPost.update_jsonb(uid, cat_dic)

    for index, catid in enumerate(new_category_arr):
        MPost2Catalog.add_record(uid, catid, index)

    # Delete the old category if not in post requests.
    for cur_info in current_infos:
        if cur_info.tag_id not in new_category_arr:
            MPost2Catalog.remove_relation(uid, cur_info.tag_id)


def fix_entity_path(sig):
    # 元数据实体目录
    ins = '/home/bk/ws/igadc_datasets'
    # ins = './entitydata'
    # copy到新目录
    out_dir = './static/datasets'

    for home, dirs, files in os.walk(ins):
        for filename in files:
            if sig in str(filename):
                entity_path = os.path.join(home, filename)
                out_path = os.path.join(out_dir, 'entityx_' + get_uu8d() + '_' + sig + '.7z')

                shutil.copy(entity_path, out_path)
                return out_path.strip('.') if out_path else None


# Todo: 内容解析错误。
def chuli_meta(sig, metafile):
    cnts = open(metafile).readlines()

    meta_dic = {}

    for row in cnts:

        if row:
            pass
        else:
            break

        the_key, the_val = [x.strip() for x in row.split(':')]

        meta_dic[the_key] = the_val


    meta_dic['identifier'] = sig

    return meta_dic


def get_meta(catid, sig, kind_sig=''):
    if kind_sig:
        pass
    else:
        return
    meta_base = './static/xx_mds'
    if os.path.exists(meta_base):
        pass
    else:
        return False

    for wroot, wdirs, wfiles in os.walk(meta_base):
        for wdir in wdirs:

            if wdir.lower().endswith(sig[1:]):
                ds_base = pathlib.Path(os.path.join(wroot, wdir))
                pp_data = {'logo': '', 'kind': kind_sig}

                for uu in ds_base.iterdir():
                    if uu.name.endswith('.md'):
                        meta_dic = chuli_meta(sig, uu)

                        if 'title' in meta_dic:
                            pass
                        else:
                            continue

                        pp_data['title'] = meta_dic['title']
                        pp_data['cnt_md'] = meta_dic['cnt_md']
                        pp_data['user_name'] = meta_dic['author']
                        pp_data['gcat0'] = catid
                        pp_data['def_cat_pid'] = catid[:2] + '00'
                        pp_data['valid'] = 1
                        # 将主要数据添加到外扩展
                        pp_data['extinfo'] = {

                        }

                        kwargsa = {
                            'gcat0': catid,
                            'cat_id': catid,
                        }
                    elif uu.name.startswith('thumbnail_'):
                        hou = os.path.splitext(uu.name)[-1]
                        if hou in ['.jpg', '.png', '.JPG', '.PNG', 'jpeg', 'JPEG']:
                            thum_path = gen_thumb(os.path.join(wroot, wdir, uu.name), 'd' + sig[1:])
                            pp_data['logo'] = thum_path.strip('.')

                # ToDo: 未处理完

                dataset_id = 'dn' + sig[1:]
                the_entity = fix_entity_path(dataset_id)

                post_id = sig

                postinfo = MPost.get_by_uid(post_id)

                if postinfo:
                    postinfo.extinfo.__update_post(pp_data['extinfo'])
                    pp_data['extinfo'] = postinfo.extinfo

                # pp_data['extinfo']['tag_data_classification'] = str(dwmode)

                if the_entity:
                    pp_data['extinfo']['tag__file_download'] = the_entity
                    MPost.update_jsonb(post_id, pp_data['extinfo'])
                else:
                    # 如果未找到实体
                    pass
                    # update_category(post_id, pp_data)

                    # ws.cell(row=i, column=2).value = pp_data['title']
                # Write a copy with titles.

                MPost.add_or_update(sig, pp_data)

                update_category(sig, pp_data, kwargsa)

                update_label(sig, pp_data)

                MPost2Catalog.add_record(sig, catid)

                return pp_data['title']


def gen_thumb(img_path, sig):
    img = Image.open(img_path)
    # newWidth = 400
    # newHeight = int(float(400) / img.size[0] * img.size[1])
    img_width = 400 if img.size[0] > 400 else img.size[0]
    img_height = 300 if img.size[1] > 300 else img.size[1]
    img.thumbnail((img_width, img_height), Image.ANTIALIAS)
    try:
        thum_file_path = os.path.join(logo_cache_dir, 'd' + sig[1:] + '.jpg')
        img.save(thum_file_path, "JPEG")
    except Exception as err:
        print(repr(err))
        thum_file_path = os.path.join(logo_cache_dir, 'd' + sig[1:] + '.png')
        img.save(thum_file_path, "PNG")
    return thum_file_path


def update_label(signature, post_data):
    '''
    Update the label .
    '''
    current_tag_infos = MPost2Label.get_by_uid(signature).objects()
    if 'tags' in post_data:
        pass
    else:
        return False

    if '；' in post_data['tags']:
        tags_arr = [x.strip() for x in post_data['tags'].split('；')]
    elif ',' in post_data['tags']:
        tags_arr = [x.strip() for x in post_data['tags'].split(',')]
    elif '，' in post_data['tags']:
        tags_arr = [x.strip() for x in post_data['tags'].split('，')]
    elif ';' in post_data['tags']:
        tags_arr = [x.strip() for x in post_data['tags'].split(';')]
    else:
        tags_arr = [x.strip() for x in post_data['tags'].split(' ')]

    if len(tags_arr) > 5:
        del tags_arr[5:]

    for tag_name in tags_arr:
        if tag_name == '':
            pass
        else:
            MPost2Label.add_record(signature, tag_name, 1)

    for cur_info in current_tag_infos:
        if cur_info.tag_name in tags_arr:
            pass
        else:
            MPost2Label.remove_relation(signature, cur_info.tag_id)


def test_import_meta():
    inws = './database/md_meta'

    for wroot, wdirs, wfiles in os.walk(inws):
        for wfile in wfiles:
            if wfile.endswith('.xlsx'):
                chli_xlsx(os.path.join(wroot, wfile))


def chli_xlsx(cat_path):
    print('操作中，请等待')
    result_xlsx = 'xx_{}'.format(os.path.split(str(cat_path))[-1])

    wb = load_workbook(cat_path)

    # uid_reg = re.compile('\_kind-\w{1}')

    # kind_pat = re.search(uid_reg, str(cat_path))
    # if kind_pat:
    kind_sig = '9'

    sheets = wb.sheetnames
    for sheet in sheets:

        catid = sheet.split('_')[0]
        ws = wb[sheet]
        rows = ws.max_row
        cols = ws.max_column
        for i in range(1, rows + 1):
            sig = ws.cell(row=i, column=1).value

            if sig:
                atitle = get_meta(catid, sig, kind_sig=kind_sig)


if __name__ == '__main__':
    test_import_meta()
