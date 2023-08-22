# -*- coding: utf-8

'''
导出数据信息
'''
import sys
import os

from torcms.model.category_model import MCategory
from torcms.model.post_model import MPost
from torcms.model.label_model import MPost2Label
from openpyxl import Workbook


def export_data():
    save_file = './static/data/'
    out_docx = os.path.join(save_file, 'xx_data_information.xlsx')
    if os.path.exists(out_docx):
        try:
            os.remove(out_docx)
        except Exception:
            pass

    wb2 = Workbook()
    ws2 = wb2['Sheet']

    ws2.cell(row=1, column=1).value = '产品编号'
    ws2.cell(row=1, column=2).value = '产品名称'
    ws2.cell(row=1, column=3).value = '简介'
    ws2.cell(row=1, column=4).value = '学科一级分类'
    ws2.cell(row=1, column=5).value = '学科分类'
    ws2.cell(row=1, column=6).value = 'SDG一级分类'
    ws2.cell(row=1, column=7).value = 'SDG分类'
    ws2.cell(row=1, column=8).value = '用户标签'

    row_inx = 2
    col_inx = 1
    recs = MPost.query_all(kind='3', limit='1000')

    for rec in recs:
        label = ''
        rec_label = MPost2Label.get_by_uid(rec.uid).objects()
        label += ','.join([r_label.tag_name for r_label in rec_label])

        infos = MPost.query_by_extinfo(key='cpbh', val=rec.extinfo['cpbh'])
        for info in infos:
            if info.kind == '3':
                xinfo = info
            if info.kind == '9':
                xinfo2 = info

        if rec:
            sub_cat_info = MCategory.get_by_uid(xinfo.extinfo['gcat0'])
            sdg_cat_info = MCategory.get_by_uid(xinfo2.extinfo['gcat0'])

            sdg_yi_category = MCategory.get_by_uid(sdg_cat_info.uid[:2] + '00')
            sub_yi_cat_info = MCategory.get_by_uid(sub_cat_info.uid[:2] + '00')

            ws2.cell(row=row_inx, column=col_inx).value = rec.extinfo['cpbh']
            ws2.cell(row=row_inx, column=col_inx + 1).value = rec.title
            ws2.cell(row=row_inx, column=col_inx + 2).value = rec.cnt_md
            ws2.cell(row=row_inx, column=col_inx + 3).value = sub_yi_cat_info.name
            ws2.cell(row=row_inx, column=col_inx + 4).value = sub_cat_info.name
            ws2.cell(row=row_inx, column=col_inx + 5).value = sdg_yi_category.name
            ws2.cell(row=row_inx, column=col_inx + 6).value = sdg_cat_info.name
            ws2.cell(row=row_inx, column=col_inx + 7).value = label

            row_inx = row_inx + 1

            wb2.save(out_docx)
            print("*" * 50)
            print("导出Data信息成功： " + save_file + 'xx_data_information.xlsx')


if __name__ == '__main__':
    export_data()
