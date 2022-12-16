# -*- coding: utf-8

'''
导入数据集的信息
'''
import sys
sys.path.append("..")
from openpyxl import load_workbook

from torcms.model.category_model import MCategory
from torcms.model.post2catalog_model import MPost2Catalog
from torcms.model.post_model import MPost
from torcms.model.label_model import MPost2Label

from pathlib import Path


import random

# jieba.analyse 用于取出字符串中的关键词。
import jieba.analyse


from torcms_taginfo.taglib.taginfo import get_tag_by_title

logo_cache_dir = './static/cache'

file_userdict = './torcms_taginfo/code_wangjy/def_userdict.txt'
jieba.load_userdict(file_userdict)

mask_arr = []
with open('./torcms_taginfo/doc/mask.txt') as fi:
    for cnt in fi.readlines():
        cnt = cnt.strip()
        if cnt:
            mask_arr.append(cnt)

the_inws = Path('./torcms_taginfo/code_wangjy')
for wfile in the_inws.rglob('def_mask*.txt'):
    print(wfile.resolve())
    with open(wfile) as fi:
        for cnt in fi.readlines():
            cnt = cnt.strip()
            if cnt:
                mask_arr.append(cnt)

mask_arr = set(mask_arr)
print(mask_arr)


def update_category(uid, post_data):
    '''
    Update the category of the post.
    :param uid:  The ID of the post. Extra info would get by requests.
    '''

    # deprecated
    # catid = kwargs['catid'] if MCategory.get_by_uid(kwargs.get('catid')) else None
    # post_data = self.get_post_data()
    if 'gcat0' in post_data:
        pass
    else:
        return False

    # Used to update MPost2Category, to keep order.
    the_cats_arr = []
    # Used to update post extinfo.
    the_cats_dict = {}


    # for old page. deprecated
    # def_cate_arr.append('def_cat_uid')

    def_cate_arr = ['gcat{0}'.format(x) for x in range(10)]
    for key in def_cate_arr:
        if key not in post_data:
            continue
        if post_data[key] == '' or post_data[key] == '0':
            continue
        # 有可能选重复了。保留前面的
        if post_data[key] in the_cats_arr:
            continue

        the_cats_arr.append(post_data[key] + ' ' * (4 - len(post_data[key])))
        the_cats_dict[key] = post_data[key] + ' ' * (4 - len(post_data[key]))

    # if catid:
    #     def_cat_id = catid
    if the_cats_arr:
        def_cat_id = the_cats_arr[0]
    else:
        def_cat_id = None

    if def_cat_id:
        the_cats_dict['gcat0'] = def_cat_id
        the_cats_dict['def_cat_uid'] = def_cat_id
        the_cats_dict['def_cat_pid'] = MCategory.get_by_uid(def_cat_id).pid



    MPost.update_jsonb(uid, the_cats_dict)

    for index, idx_catid in enumerate(the_cats_arr):
        MPost2Catalog.add_record(uid, idx_catid, index)

    # Delete the old category if not in post requests.
    current_infos = MPost2Catalog.query_by_entity_uid(uid, kind='').objects()
    for cur_info in current_infos:
        if cur_info.tag_id not in the_cats_arr:
            MPost2Catalog.remove_relation(uid, cur_info.tag_id)





def update_db_info(meta_dic, catid, sig, kind_sig):
    '''
    判断读取表格，读取相应字段后添加到数据库中。
    '''
    pp_data = {'logo': '', 'kind': kind_sig}

    pp_data['title'] = meta_dic['cpmc']

    pp_data['user_name'] = 'admin'

    pp_data['cnt_md'] = meta_dic['cp_abs']

    # pp_data['tags'] = meta_dic['tags']

    pp_data['gcat0'] = catid

    pp_data['def_cat_pid'] = catid[:2] + '00'

    # 将主要数据添加到外扩展
    pp_data['extinfo'] = {
        'cpbh': meta_dic['cpbh'],  # 产品编号
        'cplx': meta_dic['cplx'],  # 产品类型
        'xkfl': meta_dic['xkfl'],  # 学科分类
        'kjfbl': meta_dic['kjfbl'],  # 空间分辨率
        'sjfbl': meta_dic['sjfbl'],  # 时间分辨率
        'tyfs': meta_dic['tyfs'],  # 投影方式
        'dyfw': meta_dic['dyfw'],  # 地域范围
        'kssj': meta_dic['kssj'],  # 开始时间
        'jssj': meta_dic['jssj'],  # 结束时间
        'sjlx': meta_dic['sjlx'],  # 数据类型
        'sjgs': meta_dic['sjgs'],  # 数据格式
        'sjjg': meta_dic['sjjg'],  # 数据加工方法说明
        'bqsm': meta_dic['bqsm'],  # 版权声明
        'cjjg': meta_dic['cjjg'],  # 创建机构
        'cjry': meta_dic['cjry'],  # 创建人员
        'cjrq': meta_dic['cjrq'],  # 创建日期
        'fbjg': meta_dic['fbjg'],  # 发布机构
        'email': meta_dic['email'],  # 邮件
        'tel': meta_dic['tel'],  # 电话
        'zxfbrq': meta_dic['zxfbrq'],  # 最新发布日期,
        'ccl': meta_dic['ccl'],  # 存储量
        'zwjs': meta_dic['zwjs'],  # 总文件数
        'zjls': meta_dic['zjls'],  # 总记录数
        'gxfs': meta_dic['gxfs'],  # 共享方式
        'fwdz': meta_dic['fwdz']  # 访问地址
    }
    kwargsa = {
        'gcat0': catid,
        'cat_id': catid,
    }

    post_id = sig
    postinfo = MPost.get_by_uid(post_id)
    # print(pp_data)
    if postinfo:
        print(postinfo.extinfo)
        pp_data['gcat0'] = postinfo.extinfo['gcat0']

        pp_data['def_cat_pid'] = postinfo.extinfo['def_cat_pid']

        # postinfo.extinfo.update(pp_data['extinfo'])
        # pp_data['extinfo'] = postinfo.extinfo
        # # 说明内容可能会添加补充其他信息。
        # pp_data['cnt_md'] = tornado.escape.xhtml_unescape(postinfo.cnt_md)
        # # pp_data['cnt_md'] = postinfo.cnt_md
    else:
        MPost.add_or_update(sig, pp_data)

    if pp_data['title'] != '' or not pp_data['extinfo'] != '':
        MPost.add_or_update(sig, pp_data, update_time=False)

        update_category(sig, pp_data)
        update_label(sig, pp_data)
        # MPost2Catalog.add_record(sig, catid)


def update_label(signature, post_data):
    '''
    标签获取。这里判断若元数据表中存在tags字段，则生成标签并添加。若不存在则从标题中取出关键词。
    '''
    current_tag_infos = MPost2Label.get_by_uid(signature).objects()
    if 'tags' in post_data:
        tags_arr = get_tag_by_tags(post_data)
    else:
        tags_arr = get_tag_by_title(post_data)

    # print(tags_arr)

    # if len(tags_arr) > 5:
    #     del tags_arr[5:]

    for tag_name in tags_arr:
        if tag_name == '':
            pass
        else:
            MPost2Label.add_record(signature, tag_name, 1)

    for cur_info in current_tag_infos:
        if cur_info.tag_name in tags_arr:
            pass
        else:
            MPost2Label.remove_relation(signature, cur_info.tag_id)


def get_tag_by_tags(post_data):
    '''
    根据字典中tags字段，返回标签列表。
    :param post_data:
    :return:
    '''
    if '；' in post_data['tags']:
        tags_arr = [x.strip() for x in post_data['tags'].split('；')]
    elif ',' in post_data['tags']:
        tags_arr = [x.strip() for x in post_data['tags'].split(',')]
    elif '，' in post_data['tags']:
        tags_arr = [x.strip() for x in post_data['tags'].split('，')]
    elif ';' in post_data['tags']:
        tags_arr = [x.strip() for x in post_data['tags'].split(';')]
    else:
        tags_arr = [x.strip() for x in post_data['tags'].split(' ')]
    return tags_arr


def get_tag_by_title(post_data):
    '''
    根据标题返回标签列表。
    :param post_data:
    :return:
    '''
    # tags_arr = jieba.analyse.extract_tags(post_data['title'])
    raw_text = post_data['title'] + '。' + post_data['cnt_md']

    # TF-idf 关键词提取
    # tags_arr = jieba.analyse.extract_tags(raw_text)

    # textrank 关键词提取
    tags_arr = jieba.analyse.textrank(raw_text)
    print('-' * 40)
    print(tags_arr)

    out_tag = []

    for tag in tags_arr:
        if tag in mask_arr:
            pass
        else:
            out_tag.append(tag)

    return out_tag


def import_meta(kind):
    '''
    导入元数据文件夹，读取元数据。
    '''
    inw = './torcms_taginfo/code_wangjy/xx_out.xlsx'

    wb = load_workbook(inw)
    ws = wb['数据集']
    meta_dic = {}
    for row in ws.iter_rows(min_row=2):
        meta_dic['id'] = str(kind) + str(row[0].value)
        meta_dic['cpbh'] = row[1].value  # 产品编号
        meta_dic['cpmc'] = row[2].value  # 产品名称
        meta_dic['cplx'] = row[3].value  # 产品类型
        meta_dic['cp_abs'] = row[4].value  # 简介
        meta_dic['xkfl'] = row[5].value  # 学科分类
        meta_dic['tags'] = row[6].value  # 关键词
        meta_dic['kjfbl'] = row[7].value  # 空间分辨率
        meta_dic['sjfbl'] = row[8].value  # 时间分辨率
        meta_dic['tyfs'] = row[9].value  # 投影方式
        meta_dic['dyfw'] = row[10].value  # 地域范围
        meta_dic['kssj'] = row[11].value  # 开始时间
        meta_dic['jssj'] = row[12].value  # 结束时间
        meta_dic['sjlx'] = row[13].value  # 数据类型
        meta_dic['sjgs'] = row[14].value  # 数据格式
        meta_dic['sjjg'] = row[15].value  # 数据加工方法说明
        meta_dic['bqsm'] = row[16].value  # 版权声明
        meta_dic['cjjg'] = row[17].value  # 创建机构
        meta_dic['cjry'] = row[18].value  # 创建人员
        meta_dic['cjrq'] = row[19].value  # 创建日期
        meta_dic['fbjg'] = row[20].value  # 发布机构
        meta_dic['email'] = row[21].value  # 邮件
        meta_dic['tel'] = row[22].value  # 电话
        meta_dic['zxfbrq'] = row[23].value  # 最新发布日期
        meta_dic['ccl'] = row[24].value  # 存储量
        meta_dic['zwjs'] = row[25].value  # 总文件数
        meta_dic['zjls'] = row[26].value  # 总记录数
        meta_dic['gxfs'] = row[27].value  # 共享方式
        meta_dic['fwdz'] = row[28].value  # 访问地址

        sig = meta_dic['id']
        catid = get_catid(meta_dic['xkfl'], kind)
        kind = catid[:1]

        update_db_info(meta_dic, catid, sig, kind)


'''
    list = []
    for kind in kind_data:
        catlists = MCategory.query_all(kind=kind)
        for x in catlists:
            if x not in list: pass
            list.append(x)
    # 随即从列表中取出一个id值。转换成字符串方便后面调用。
    catid = str(random.choice(list))

    return catid
'''


def get_catid(cat_name, kind):
    '''
    随机读取一个分类id。默认设定分类id的第一个字符为此分类的kind值。
    :return:
    '''
    list = []

    catlists = MCategory.query_all(kind=kind)
    for x in catlists:
        if x not in list: pass
        list.append(x)

    # 随即从列表中取出一个id值。转换成字符串方便后面调用。
    catid = str(random.choice(list))
    return catid


if __name__ == '__main__':
    args = sys.argv
    kind = '3'
    import_meta(kind)
    kind = '9'
    import_meta(kind)
