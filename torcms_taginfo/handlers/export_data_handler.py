# -*- coding:utf-8 -*-

'''
导出数据
'''

import json
import os
from torcms.core.base_handler import BaseHandler
from torcms.model.category_model import MCategory
from torcms.model.post_model import MPost
from openpyxl import Workbook
from torcms.model.label_model import MPost2Label


class DownloadDataHandler(BaseHandler):
    def initialize(self, **kwargs):
        super().initialize()
        self.kind = '3'
        self.filter_view = True

    def get(self, *args, **kwargs):
        url_str = args[0]
        url_arr = self.parse_url(url_str)

        if url_str == '' or url_str == 'list':
            self.list()

        else:
            kwd = {
                'info': 'The Page not Found.',
            }
            self.show404(kwd=kwd)

    def list(self):
        # save_file = './static/data/'
        # out_docx = os.path.join(save_file, 'xx_data_information.xlsx')
        # if os.path.exists(out_docx):
        #     try:
        #         os.remove(out_docx)
        #     except Exception:
        #         pass
        #
        # wb2 = Workbook()
        # ws2 = wb2['Sheet']
        #
        # ws2.cell(row=1, column=1).value = '产品编号'
        # ws2.cell(row=1, column=2).value = '产品名称'
        # ws2.cell(row=1, column=3).value = '简介'
        # ws2.cell(row=1, column=4).value = '学科一级分类'
        # ws2.cell(row=1, column=5).value = '学科分类'
        # ws2.cell(row=1, column=6).value = 'SDG一级分类'
        # ws2.cell(row=1, column=7).value = 'SDG分类'
        # ws2.cell(row=1, column=8).value = '用户标签'
        #
        # row_inx = 2
        # col_inx = 1
        # recs = MPost.query_all(kind='n', limit='1000')
        #
        # for rec in recs:
        #     label = ''
        #     rec_label = MPost2Label.get_by_uid(rec.uid).objects()
        #     for r_label in rec_label:
        #         label += r_label.tag_name + ','
        #
        #     infos = MPost.query_by_extinfo(key='cpbh', val=rec.extinfo['cpbh'])
        #     for info in infos:
        #         if info.kind == 'n':
        #             xinfo = info
        #         if info.kind == 's':
        #             xinfo2 = info
        #
        #     if rec:
        #         sub_cat_info = MCategory.get_by_uid(xinfo.extinfo['gcat0'])
        #         sdg_cat_info = MCategory.get_by_uid(xinfo2.extinfo['gcat0'])
        #
        #         sdg_yi_category = MCategory.get_by_uid(sdg_cat_info.uid[:2] + '00')
        #         sub_yi_cat_info = MCategory.get_by_uid(sub_cat_info.uid[:2] + '00')
        #
        #         ws2.cell(row=row_inx, column=col_inx).value = rec.extinfo['cpbh']
        #         ws2.cell(row=row_inx, column=col_inx + 1).value = rec.title
        #         ws2.cell(row=row_inx, column=col_inx + 2).value = rec.cnt_md
        #         ws2.cell(row=row_inx, column=col_inx + 3).value = sub_yi_cat_info.name
        #         ws2.cell(row=row_inx, column=col_inx + 4).value = sub_cat_info.name
        #         ws2.cell(row=row_inx, column=col_inx + 5).value = sdg_yi_category.name
        #         ws2.cell(row=row_inx, column=col_inx + 6).value = sdg_cat_info.name
        #         ws2.cell(row=row_inx, column=col_inx + 7).value = label
        #
        #         row_inx = row_inx + 1
        #
        #         wb2.save(out_docx)
        #         print("*" * 50)
        #         print("导出Data信息成功： " + save_file + 'xx_data_information.xlsx')

        out_docx= '/static/data/xx_data_information.xlsx'
        self.render('../torcms_taginfo/download_data.html',
                    userinfo=self.userinfo,
                    out_docx=out_docx,
                    kwd={})
