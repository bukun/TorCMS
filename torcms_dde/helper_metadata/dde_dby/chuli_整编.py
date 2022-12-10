from pathlib import Path
from osgeo import ogr
import fiona
from pyproj import Proj
from pyproj import CRS
from osgeo import osr

target = osr.SpatialReference()
target.ImportFromEPSG(4326)


def get_geo(wdir):
    a, b, c, d = -180, -90, 180, 90
    sig = False
    for wfile in wdir.rglob('*.shp'):
        sig = True
        print('    ', wfile)
        # ds = ogr.Open(str(wfile))
        # lyr = ds.GetLayer(0)
        # ext = lyr.GetExtent()
        # env = lyr.GetEnvelope()
        try:
            ds = fiona.open(wfile)
        except:
            continue

        # print(ds.crs)

        try:
            crs = CRS(ds.crs)
            # print('=' * 20)
            # print(type(crs))
            # print(crs)
            # print(type(target))
            # print(target)
        except:
            continue

        p = Proj(crs)
        # print(p)

        ll = p(ds.bounds[0], ds.bounds[1], inverse=True)
        ur = p(ds.bounds[2], ds.bounds[3], inverse=True)
        # srcp = osr.SpatialReference()
        # srcp.ImportFromEPSG(int(ds.crs['init'].split(':')[-1]))
        # trans = osr.CoordinateTransformation(srcp, target)
        # ll = trans.TransformPoint(ds.bounds[0], ds.bounds[1])
        # ur = trans.TransformPoint(ds.bounds[2], ds.bounds[3])

        if ll[0] > 90 or ll[1] > 90 or ur[0] > 90 or ur[1] > 90:
            print(ds.bounds)
            print(ll, ur)

        if ll[0] > a:
            a = ll[0]
        if ll[1] > b:
            b = ll[1]
        if ur[0] < c:
            c = ur[0]
        if ur[1] < d:
            d = ur[1]

    if sig:
        return a, b, c, d
    else:
        return None

        # print(ext)
        # print(env)


from pathlib import Path
from openpyxl import load_workbook
from xml.sax.saxutils import escape

'''
导入数据集的元数据，以及数据实体
'''

import os
import pathlib
from openpyxl import load_workbook


# HOST_QGSVR = '39.100.254.142:6626'


def chuli_meta(metafile, sig, ginfo=None):
    print(metafile)
    wb = load_workbook(str(metafile))
    sheet = wb[wb.sheetnames[0]]

    outws = Path('./xx_xml')

    if outws.exists():
        pass
    else:
        outws.mkdir()

    tmpl = open('./tmpl.xml').read()

    tmpl_0 = '''<?xml version="1.0" encoding="UTF-8"?>
   <csw:Record
    xmlns:csw="http://www.opengis.net/cat/csw/2.0.2"
    xmlns:dc="http://purl.org/dc/elements/1.1/"
    xmlns:dct="http://purl.org/dc/terms/"
    xmlns:ows="http://www.opengis.net/ows"
    xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance"
    xsi:schemaLocation="http://www.opengis.net/cat/csw/2.0.2/record.xsd">
           '''
    tmpl_geo = '''
<ows:BoundingBox crs="urn:x-ogc:def:crs:EPSG:6.11:4326">
<ows:LowerCorner>{} {}</ows:LowerCorner>
<ows:UpperCorner>{} {}</ows:UpperCorner>
</ows:BoundingBox>
    '''

    tp_identifier = '<dc:identifier>{}</dc:identifier>'

    tmpl_9 = ''' 
     </csw:Record>
                         
        '''

    outfile = outws / (sig + '.xml')

    with open(outfile, 'w') as fo:
        fo.write(tmpl_0)

        for row in sheet.iter_rows(min_row=2):
            if not (row[1].value):
                continue
            else:
                pass

            # print(row[0].value.strip())
            if row[0].value.strip() in ['creator', 'contributer', 'subject', 'relation', 'title', 'type', 'source']:
                fo.write('<dc:{0}>{1}</dc:{0}>\n'.format(row[0].value, escape(row[1].value)))
            if row[0].value.strip() in ['abstract']:
                fo.write('<dct:{0}>{1}</dct:{0}>\n'.format(row[0].value, escape(row[1].value)))
        fo.write(tp_identifier.format('drrmd-' + sig))
        if ginfo:
            fo.write(tmpl_geo.format(ginfo[0], ginfo[1], ginfo[2], ginfo[3]))
            # fo.write(tmpl_geo.format(ginfo[1], ginfo[0], ginfo[3], ginfo[2]))
        fo.write(tmpl_9)


def get_meta():
    '''
    Get metadata of dataset via ID.
    '''

    meta_base = '/home/bk/geows'

    # meta_base = './dataset_DDE'
    #
    # if os.path.exists(meta_base):
    #     pass
    # else:
    #     return False
    #
    for wroot, wdirs, wfiles in os.walk(meta_base):
        # print(wdirs)
        for wdir in wdirs:
            if wdir.startswith('dataset') and '_dn' in wdir:

                #  Got the dataset of certain ID.

                # ds_base = pathlib.Path()
                ds_base = pathlib.Path(os.path.join(wroot, wdir))

                geoinfo = get_geo(ds_base)

                dataset_id = ds_base.name.split('_')[-1]

                for uu in ds_base.iterdir():

                    if uu.name.endswith('.xlsx'):
                        if uu.name.startswith('~') or uu.name.startswith('.~'):
                            pass
                        else:
                            chuli_meta(uu, dataset_id, ginfo=geoinfo)


if __name__ == '__main__':
    get_meta()
