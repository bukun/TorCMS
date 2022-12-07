from pathlib import Path
from openpyxl import load_workbook
from xml.sax.saxutils import escape

'''
导入数据集的元数据，以及数据实体
'''

import os
import pathlib
from openpyxl import load_workbook

HOST_QGSVR = '39.100.254.142:6626'


def chuli_meta(metafile, sig):
    wb = load_workbook(str(metafile))
    sheet = wb[wb.sheetnames[0]]

    outws = Path('./xx_xml')

    if outws.exists():
        pass
    else:
        outws.mkdir()

    tmpl = open('./tmpl.xml').read()

    tmpl_0 = '''<?xml version="1.0" encoding="UTF-8"?>
   <csw:Record
    xmlns:csw="http://www.opengis.net/cat/csw/2.0.2"
    xmlns:dc="http://purl.org/dc/elements/1.1/"
    xmlns:dct="http://purl.org/dc/terms/"
    xmlns:ows="http://www.opengis.net/ows"
    xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance"
    xsi:schemaLocation="http://www.opengis.net/cat/csw/2.0.2/record.xsd">
           '''

    tp_identifier = '<dc:identifier>{}</dc:identifier>'

    tmpl_9 = ''' 
     </csw:Record>
                         
        '''

    outfile = outws / (sig + '.xml')

    with open(outfile, 'w') as fo:
        fo.write(tmpl_0)

        for row in sheet.iter_rows(min_row=2):
            if not (row[1].value):
                continue
            else:
                pass

            print(row[0].value.strip())
            if row[0].value.strip() in ['creator', 'contributer', 'subject', 'relation', 'title', 'type']:
                fo.write('<dc:{0}>{1}</dc:{0}>\n'.format(row[0].value, escape(row[1].value)))
            if row[0].value.strip() in ['abstract']:
                fo.write('<dct:{0}>{1}</dct:{0}>\n'.format(row[0].value, escape(row[1].value)))
        fo.write(tp_identifier.format('drrks2022-' + sig))
        fo.write(tmpl_9)


def get_meta():
    '''
    Get metadata of dataset via ID.
    '''

    meta_base = './xx_it'
    # meta_base = './dataset_DDE'
    #
    # if os.path.exists(meta_base):
    #     pass
    # else:
    #     return False
    #
    for wroot, wdirs, wfiles in os.walk(meta_base):
        # print(wdirs)
        for wdir in wdirs:
            if wdir.startswith('dataset') and '_dn' in wdir:

                #  Got the dataset of certain ID.

                # ds_base = pathlib.Path()
                ds_base = pathlib.Path(os.path.join(wroot, wdir))

                dataset_id = ds_base.name.split('_')[-1]
                sig = '9' + str(dataset_id[2:])

                for uu in ds_base.iterdir():

                    if uu.name.endswith('.xlsx') and not uu.name.startswith('~'):
                        chuli_meta(uu, sig)


# def import_meta():
#
#
#     for wroot, wdirs, wfiles in os.walk(inws):
#
#         for wdir in wdirs:
#
#             if len(wdir.split('_')) >= 2:
#                 pass
#             else:
#                 continue
#             if wdir.split('_')[2][:3] == 'pid':
#                 cat_id = wdir.split('_')[2][3:]
#
#                 if wdir.lower().endswith("pid" + cat_id):
#                     #  Got the dataset of certain ID.
#                     ds_base = pathlib.Path(os.path.join(wroot, wdir))
#                     for uu in ds_base.iterdir():
#                         print(uu)
#
#                         catid = '1212'
#
#                         get_meta(catid, uu)


if __name__ == '__main__':
    get_meta()
