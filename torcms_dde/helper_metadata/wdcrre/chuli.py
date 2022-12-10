from html.parser import HTMLParser
import tornado.escape
from re import sub
from sys import stderr
from traceback import print_exc
import sys
from pathlib import Path
from openpyxl import load_workbook
from xml.sax.saxutils import escape

inws = Path('./esheet')

outws = Path('./xx_xml')

if outws.exists():
    pass
else:
    outws.mkdir()

tmpl = open('./tmpl.xml').read()

# print(tmpl)

tmpl_0 = '''<?xml version="1.0" encoding="UTF-8"?>
<csw:Record
	xmlns:csw="http://www.opengis.net/cat/csw/2.0.2"
	xmlns:dc="http://purl.org/dc/elements/1.1/"
	xmlns:dct="http://purl.org/dc/terms/"
	xmlns:ows="http://www.opengis.net/ows"
	xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance"
	xsi:schemaLocation="http://www.opengis.net/cat/csw/2.0.2/record.xsd">
'''

tmpl_9 = '''
</csw:Record>
'''

tp_creator = '<dc:creator>{}</dc:creator>'
tp_contributor = '<dc:contributor>{}</dc:contributor>'
tp_publisher = '<dc:publisher>{}</dc:publisher>'
tp_subject = '<dc:subject>{}</dc:subject>'
tp_abstract = '<dct:abstract>{}</dct:abstract>'
tp_identifier = '<dc:identifier>{}</dc:identifier>'
tp_relation = '<dc:relation>{}</dc:relation>'
tp_source = '<dc:source>{}</dc:source>'
tp_rights = '<dc:rights>{}</dc:rights>'
tp_type = '<dc:type>{}</dc:type>'
tp_title = '<dc:title>{}</dc:title>'
tp_mofified = '<dct:modified>{}</dct:modified>'
tp_language = '<dc:language>{}</dc:language>'

# sigs = ['a', 'b', 'c','d','e','f','g','h', 'i', 'j', 'k', 'l', 'm', 'n', 'o','p','q']


def chuli_cell(cell):

    cv = cell.value if cell.value else ''
    return escape(cv )

for xlsx_file in inws.rglob('*.xlsx'):
    if xlsx_file.name.startswith('~$'):
        continue
    try:
        print(xlsx_file.name)
        wb = load_workbook(xlsx_file)
        del(wb)
    except:
        print('Error:', xlsx_file.name)
        sys.exit()


idx = 0


xlsx_file =  './esheet/wdcrre_data.xlsx'

# dir_sig = xlsx_file.stem

# sig = f'xh_{idx}_'
#
# idx = idx + 1




class _DeHTMLParser(HTMLParser):
    def __init__(self):
        HTMLParser.__init__(self)
        self.__text = []

    def handle_data(self, data):
        text = data.strip()
        if len(text) > 0:
            text = sub('[ \t\r\n]+', ' ', text)
            self.__text.append(text + ' ')

    def handle_starttag(self, tag, attrs):
        if tag == 'p':
            self.__text.append('\n\n')
        elif tag == 'br':
            self.__text.append('\n')

    def handle_startendtag(self, tag, attrs):
        if tag == 'br':
            self.__text.append('\n\n')

    def text(self):
        return ''.join(self.__text).strip()


def dehtml(text):
    try:
        parser = _DeHTMLParser()
        parser.feed(text)
        parser.close()
        return parser.text()
    except:
        print_exc(file=stderr)
        return text



wb = load_workbook(xlsx_file)

ws = wb.active
print(ws.title)

for row in ws.rows:
    if row[0]:
        pass
    else:
        break
    # print(row)

    v0 = chuli_cell(row[0])  #id
    v1 = chuli_cell(row[1]) # title
    v2 = chuli_cell(row[2])     # keywords
    v3 = chuli_cell(row[3])     # date
    v4 = chuli_cell(row[4])     # cnt_html



    outfile = outws / f'wdcrre-{v0}.xml'
    with open(outfile, 'w') as fo:

        fo.write(tmpl_0)
        fo.write('\n')
        fo.write(tp_identifier.format(f'wdcrre-{v0}'))
        fo.write('\n')
        fo.write(tp_language.format('English'))
        fo.write('\n')
        # fo.write(tp_subject.format(v2))
        fo.write('\n')
        fo.write(tp_title.format(v1))
        fo.write('\n')
        fo.write(tp_source.format(f'http://wdcrre.data.ac.cn/meta_info/{v0}'))
        fo.write('\n')
        # fo.write(tp_relation.format(v5))
        fo.write('\n')
        fo.write(
            tp_abstract.format(
                dehtml(
                    tornado.escape.xhtml_unescape(
                    tornado.escape.xhtml_unescape(
                        v4
                    )
                    )
                )
            )
        )
        # fo.write(tp_contributor.format(v9))

        fo.write(tmpl_9)

