import tornado.escape
import json
import config
import os
import uuid
import openpyxl
import openpyxl.styles
from config import CMS_CFG, router_post
from torcms.handlers.post_handler import PostHandler
from torcms.core.tools import logger
from torcms.model.post_model import MPost
from openpyxl import load_workbook
from torcms.core import privilege
from torcms.model.category_model import MCategory
from torcms.model.post2catalog_model import MPost2Catalog
from torcms.model.label_model import MPost2Label
from torcms.model.post_hist_model import MPostHist
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from torcms.model.entity_model import MEntity


def update_category(uid, post_data):
    '''
    Update the category of the post.
    :param uid:  The ID of the post. Extra info would get by requests.
    '''

    # deprecated
    # catid = kwargs['catid'] if MCategory.get_by_uid(kwargs.get('catid')) else None
    # post_data = self.get_request_arguments()
    if 'gcat0' in post_data:
        pass
    else:
        return False

    # Used to update MPost2Category, to keep order.
    the_cats_arr = []
    # Used to update post extinfo.
    the_cats_dict = {}

    # for old page. deprecated
    # def_cate_arr.append('def_cat_uid')

    def_cate_arr = ['gcat{0}'.format(x) for x in range(10)]
    for key in def_cate_arr:
        if key not in post_data:
            continue
        if post_data[key] == '' or post_data[key] == '0':
            continue
        # 有可能选重复了。保留前面的
        if post_data[key] in the_cats_arr:
            continue

        the_cats_arr.append(post_data[key] + ' ' * (4 - len(post_data[key])))
        the_cats_dict[key] = post_data[key] + ' ' * (4 - len(post_data[key]))

    # if catid:
    #     def_cat_id = catid
    if the_cats_arr:
        def_cat_id = the_cats_arr[0]
    else:
        def_cat_id = None

    if def_cat_id:
        the_cats_dict['gcat0'] = def_cat_id
        the_cats_dict['def_cat_uid'] = def_cat_id
        the_cats_dict['def_cat_pid'] = MCategory.get_by_uid(def_cat_id).pid

    # Add the category
    logger.info('Update category: {0}'.format(the_cats_arr))
    logger.info('Update category: {0}'.format(the_cats_dict))

    MPost.update_jsonb(uid, the_cats_dict)

    for index, idx_catid in enumerate(the_cats_arr):
        MPost2Catalog.add_record(uid, idx_catid, index)

    # Delete the old category if not in post requests.
    current_infos = MPost2Catalog.query_by_entity_uid(uid, kind='').objects()
    for cur_info in current_infos:
        if cur_info.tag_id not in the_cats_arr:
            MPost2Catalog.remove_relation(uid, cur_info.tag_id)


def update_label(signature, post_data):
    '''
    Update the label when updating.
    '''
    current_tag_infos = MPost2Label.get_by_uid(signature).objects()
    if 'tags' in post_data:
        pass
    else:
        return False
    if '；' in post_data['tags']:
        tags_arr = [x.strip() for x in post_data['tags'].split('；')]
    elif ',' in post_data['tags']:
        tags_arr = [x.strip() for x in post_data['tags'].split(',')]
    elif '，' in post_data['tags']:
        tags_arr = [x.strip() for x in post_data['tags'].split('，')]
    elif ';' in post_data['tags']:
        tags_arr = [x.strip() for x in post_data['tags'].split(';')]
    else:
        tags_arr = [x.strip() for x in post_data['tags'].split(' ')]

    for tag_name in tags_arr:
        if tag_name == '':
            pass
        else:
            MPost2Label.add_record(signature, tag_name, 1)

    for cur_info in current_tag_infos:
        if cur_info.tag_name in tags_arr:
            pass
        else:
            MPost2Label.remove_relation(signature, cur_info.tag_id)


class MetadataHandler(PostHandler):

    def initialize(self, **kwargs):
        super(MetadataHandler, self).initialize(**kwargs)

        self.kind = kwargs.get('kind', 'd')
        self.filter_view = kwargs.get('filter_view', True)

    def get(self, *args, **kwargs):
        url_str = args[0]
        url_arr = self.parse_url(url_str)
        if url_str in ['data', 'index', '']:
            self.index()
        elif url_arr[0] == '_cat_add':
            self._to_add(catid=url_arr[1])

        elif url_arr[0] == 'download_excel':

            self.download_xlsx(url_arr[1])

        elif url_arr[0] == '_add':
            if len(url_arr) == 2:
                self._to_add(uid=url_arr[1])
            else:
                self._to_add()
        elif len(url_arr) == 1:
            self._view_or_add(url_str)
        elif len(url_arr) == 2:
            dict_get = {
                '_edit_kind': self._to_edit_kind,
                '_edit': self._to_edit,
                '_delete': self._delete,
            }
            dict_get.get(url_arr[0])(url_arr[1])
        else:
            self.show404()

    def post(self, *args, **kwargs):

        url_str = args[0]
        logger.info('Post url: {0}'.format(url_str))
        url_arr = self.parse_url(url_str)

        if url_arr[0] in ['_edit']:
            self.update(url_arr[1])

        elif url_arr[0] in ['_add']:
            if len(url_arr) == 2:
                self.add(uid=url_arr[1])
            else:
                self.add()
        elif url_arr[0] == 'upload_excel':
            self.upload_excel()
        elif url_arr[0] == '_edit_kind':
            self._change_kind(url_arr[1])
        elif url_arr[0] in ['_cat_add']:
            self.add(catid=url_arr[1])
        elif len(url_arr) == 1 and len(url_str) >= 4:
            self.add(uid=url_str)
        elif url_arr[0] == 'rel' and len(url_arr) == 3:
            self._add_relation(url_arr[1], url_arr[2])

        else:
            self.show404()


    def chuli_meta(self, metafile):
        try:
            wb = load_workbook(str(metafile))
        except:
            return None
        sheet = wb[wb.sheetnames[0]]
        meta_dic = {}

        ii = 1
        for row in sheet.iter_rows(min_row=2):

            if not (row[1].value) or sheet.cell(2, 1).value != 'identifier':
                continue
            else:
                pass
            ii = ii + 1
            if row[0].value in ['date', 'date_revision', 'date_creation', 'date_publication', 'date_modified',
                                'insert_date', 'time_begin', 'time_end']:
                meta_dic['pycsw_' + str(row[0].value).strip()] = str(row[1].value)[:10]
            else:
                meta_dic['pycsw_' + str(row[0].value).strip()] = str(row[1].value)

        return meta_dic

    def fetch_post_data(self):
        '''
        fetch post accessed data. post_data, and ext_dic.
        '''
        post_data = {}
        ext_dic = {}
        ii = 1
        for key in self.request.arguments:
            if key.startswith('ext_') or key.startswith('tag_') or key.startswith('pycsw_'):
                if key.startswith('tag_meta_'):
                    if self.get_argument(key) != '':
                        ext_dic[key] = {
                            'label': str(key)[13:],
                            'info': self.get_argument(key, default=''),
                            'order': ii,
                        }

                        ii = ii + 1
                else:
                    ext_dic[key] = self.get_argument(key, default='')

            else:
                post_data[key] = self.get_arguments(key)[0]

        post_data['user_name'] = self.userinfo.user_name
        post_data['kind'] = self.kind

        if 'tags' in post_data:
            ext_dic['def_tag_arr'] = [
                x.strip() for x in post_data['tags'].strip().strip(',').split(',')
            ]
        ext_dic = dict(ext_dic, **self.ext_post_data(postdata=post_data))

        return (post_data, ext_dic)

    @tornado.web.authenticated
    @privilege.auth_add
    # @tornado.web.asynchronous
    @tornado.gen.coroutine
    def add(self, **kwargs):
        '''
        in infor.
        '''

        post_data, ext_dic = self.fetch_post_data()

        if ext_dic.get('pycsw_identifier'):
            uid = ext_dic.get('pycsw_identifier')
        elif post_data.get('uid'):
            uid = post_data['uid']
        else:
            uid = self._gen_uid()

        title = post_data['title'].strip()

        if len(title) < 2:
            kwd = {
                'info': 'Title cannot be less than 2 characters',
                'link': '/'
            }
            self.render('misc/html/404.html', userinfo=self.userinfo, kwd=kwd)

        if 'gcat0' in post_data:
            pass
        else:
            return False

        if 'valid' in post_data:
            post_data['valid'] = int(post_data['valid'])
        else:
            post_data['valid'] = 1

        ext_dic['def_uid'] = uid
        ext_dic['gcat0'] = post_data['gcat0']
        ext_dic['def_cat_uid'] = post_data['gcat0']
        ext_dic['status'] = 'a0'
        # -------------重複值處理--------------------------
        ext_dic['tag__language'] = ext_dic['pycsw_language']
        ext_dic['tag__data_type'] = ext_dic['pycsw_type']
        if ext_dic.get('pycsw_format'):
            format_dic = {'raster': '1', 'vector': '2', 'imagery': '3', 'esheet': '4', 'database': '5', 'file': '6',
                          'excel': '7', 'tif': '8', 'shp': '9', 'grd': '10', 'geotiff': '11'}
            for format_key, format_value in format_dic.items():
                if ext_dic.get('pycsw_format').lower() == format_key:
                    ext_dic['tag_data_format'] = format_value

        # -----------------------------------------------

        MPost.add_or_update_post(ext_dic['def_uid'], post_data, extinfo=ext_dic)
        kwargs.pop('uid', None)  # delete `uid` if exists in kwargs

        self._add_download_entity(ext_dic)

        update_category(ext_dic['def_uid'], post_data)
        update_label(ext_dic['def_uid'], post_data)
        if ext_dic.get('pycsw_file'):
            file_path = ext_dic['pycsw_file']
            file_src = os.path.join('.' + file_path)
            try:
                wb = load_workbook(str(file_src))
            except:
                return None
            sheet = wb[wb.sheetnames[0]]
            sheet.cell(2, 2).value = uid
            wb.save(file_src)

        tornado.ioloop.IOLoop.instance().add_callback(self.cele_gen_whoosh)
        self.redirect('/{0}/{1}'.format(router_post[self.kind], uid))

    @tornado.web.authenticated
    @privilege.auth_edit
    @tornado.gen.coroutine
    def update(self, uid):
        '''
        in infor.
        '''

        post_data, ext_dic = self.fetch_post_data()
        if 'gcat0' in post_data:
            pass
        else:
            return False

        if ext_dic.get('pycsw_identifier'):
            postinfo = MPost.get_by_uid(ext_dic.get('pycsw_identifier'))
            if postinfo:
                postinfo = MPost.get_by_uid(ext_dic.get('pycsw_identifier'))
            else:
                return self.add()
        else:
            postinfo = MPost.get_by_uid(uid)

        if 'valid' in post_data:
            post_data['valid'] = int(post_data['valid'])
        else:
            post_data['valid'] = postinfo.valid

        ext_dic['def_uid'] = postinfo.uid
        # 用于判断是第几轮审核
        if postinfo.state[1] == '3':
            edit_count = int(postinfo.extinfo.get('def_edit_count', 0)) + 1
        else:
            edit_count = int(postinfo.extinfo.get('def_edit_count', 0))

        ext_dic['def_edit_count'] = edit_count
        # 用于记录审核通过后第几轮修改
        if postinfo.state[1] == '2':
            approved_count = int(postinfo.extinfo.get('def_approved_count', 0)) + 1
        else:
            approved_count = int(postinfo.extinfo.get('def_approved_count', 0))

        ext_dic['def_approved_count'] = approved_count

        # -------------重複值處理--------------------------
        ext_dic['tag__language'] = ext_dic['pycsw_language']
        ext_dic['tag__data_type'] = ext_dic['pycsw_type']
        if ext_dic.get('pycsw_format'):
            format_dic = {'raster': '1', 'vector': '2', 'imagery': '3', 'esheet': '4', 'database': '5', 'file': '6',
                          'excel': '7', 'tif': '8', 'shp': '9', 'grd': '10', 'geotiff': '11'}
            for format_key, format_value in format_dic.items():
                if ext_dic.get('pycsw_format').lower() == format_key:
                    ext_dic['tag_data_format'] = format_value

        # -----------------------------------------------

        cnt_old = tornado.escape.xhtml_unescape(postinfo.cnt_md).strip()
        cnt_new = post_data['cnt_md'].strip()
        if cnt_old == cnt_new:
            pass
        else:
            MPostHist.create_post_history(postinfo, self.userinfo)

        MPost.add_or_update_post(postinfo.uid, post_data, extinfo=ext_dic)
        # todo:应该判断当前审核状态，是否可以进行修改状态。
        if postinfo.state[1] == '3':
            state_arr = ['a', 'b', 'c', 'd', 'e', 'f']
            if len(state_arr) > edit_count:
                pstate = str(state_arr[edit_count]) + '0' + postinfo.state[2:]

                MPost.update_state(uid, pstate)

        if postinfo.state[1] == '2':
            approved_state_arr = ['1', '2', '3', '4', '5', '6', '7', '8', '9']
            if len(approved_state_arr) > approved_count:
                pstate = postinfo.state[0] + '1' + str(approved_state_arr[approved_count]) + '0'
            else:
                pstate = postinfo.state[0] + '190'
            MPost.update_state(uid, pstate)
        self._add_download_entity(ext_dic)

        update_category(postinfo.uid, post_data)
        update_label(postinfo.uid, post_data)
        if ext_dic.get('pycsw_file') not in ['', 'None']:
            file_path = ext_dic['pycsw_file']
            file_src = os.path.join('.' + file_path)
            try:
                wb = load_workbook(str(file_src))
            except:
                return None
            sheet = wb[wb.sheetnames[0]]
            sheet.cell(2, 2).value = postinfo.uid
            wb.save(file_src)

        logger.info('post kind:' + self.kind)
        tornado.ioloop.IOLoop.instance().add_callback(self.cele_gen_whoosh)
        self.redirect('/{0}/{1}'.format(router_post[postinfo.kind], postinfo.uid))

    def download_xlsx(self, postid):
        '''
        下载元数据XLSX文件
        '''
        tag_info = MPost2Label.get_by_uid(postid).objects()
        print('======')
        keywords = ''
        for x in tag_info:
            if x.tag_name not in keywords: keywords = keywords + str(x.tag_name) + ','

        file_src2 = os.path.join('./torcms_metadata/meta_元数据模板20220921.xlsx')
        tmp_file = './static/xx_metadata_{0}.xlsx'.format(postid)
        postinfo = MPost.get_by_uid(postid)
        if os.path.exists(tmp_file):
            os.remove(tmp_file)
        try:
            wb = load_workbook(file_src2)
        except:
            return None

        ws = wb.active

        validator_lang = self.生成验证语言()

        # c1 = ws["B7"]
        # c1.value = "英文"
        # validator_lang.add(c1)

        # Add the data-validation object to the worksheet
        ws.add_data_validation(validator_lang)
        ws.cell(2, 2).value = postinfo.uid
        ws.cell(3, 2).value = postinfo.title
        for ii in range(4, ws.max_row + 1):
            if ii == 7:
                pass
            if str(ws.cell(row=ii, column=1).value).strip() == 'keywords':
                post_value = str(keywords)
            else:
                meta_name = 'pycsw_' + str(ws.cell(row=ii, column=1).value).strip()
                print(meta_name)

                post_value = postinfo.extinfo[meta_name] if meta_name in postinfo.extinfo else ''
            # print(post_value)
            ws.cell(ii, 2).value = post_value
        ws.cell(2, 2).protection = openpyxl.styles.Protection(locked=True)

        ws.protection.sheet = True
        ws.protection.enable()
        wb.save(tmp_file)
        wb.close()

        output = {f'file_src': tmp_file.strip('.')}

        return json.dump(output, self)

    def 生成验证语言(self):
        dv = DataValidation(type="list", formula1='"英文,中文,法文,俄语,阿拉伯语,西班牙语,其他"', allow_blank=True)
        dv.error = 'Your entry is not in the list'
        dv.errorTitle = 'Invalid Entry'
        dv.prompt = 'Please select from the list'
        dv.promptTitle = 'List Selection'
        return dv

    @tornado.web.authenticated
    def upload_excel(self):
        '''
        Adding the pdf file.
        '''

        post_data = self.get_request_arguments()

        img_entity = self.request.files['file'][0]
        img_desc = img_entity["filename"]
        filename = img_entity["filename"]

        if filename and self.allowed_file_pdf(filename):
            pass
        else:
            kwd = {
                'pager': '',
                'err_info': '* The formats of uploadable files are: xlsx'
            }
            self.render('misc/entity/entity_add.html',
                        cfg=config.CMS_CFG,
                        kwd=kwd,
                        userinfo=self.userinfo)
            # return False
        if 'uid' in post_data:
            uid = post_data['uid']
        else:
            uid = self._gen_uid()
        _, hou = os.path.splitext(filename)
        signature = str(uuid.uuid1())
        outfilename = 'metadata_{0}{1}'.format(uid, hou)
        outpath = 'static/upload'
        if os.path.exists(outpath):
            pass
        else:
            os.makedirs(outpath)
        with open(os.path.join(outpath, outfilename), "wb") as fout:
            fout.write(img_entity["body"])

        sig_save = os.path.join(signature[:2], signature)
        path_save = os.path.join(signature[:2], outfilename)
        ispdf = MEntity.get_id_by_impath(outfilename)
        if ispdf:
            pass
        else:
            create_pdf = MEntity.create_entity(
                signature,
                outfilename,
                img_desc,
                kind=post_data['kind'] if 'kind' in post_data else '2')

        ext_dic = {}
        file_src = os.path.join('./static/upload/' + outfilename)

        meta_dics = self.chuli_meta(file_src)
        if meta_dics == '':
            ext_dic = {}
        else:
            for meta_dic in meta_dics:
                ext_dic[meta_dic] = meta_dics[meta_dic]
            # ext_dic['logo'] = ''
            ext_dic['kind'] = self.kind

            ext_dic['title'] = meta_dics.get('pycsw_title')
            ext_dic['cnt_md'] = meta_dics.get('pycsw_abstract')
            ext_dic['user_name'] = self.userinfo.user_name
            ext_dic['tags'] = meta_dics.get('pycsw_keywords')

            ext_dic['pycsw_file'] = '/static/upload/' + outfilename

        return json.dump(ext_dic, self)

    def allowed_file_pdf(self, filename):
        '''
        Allowed xlsx files
        '''
        ALLOWED_EXTENSIONS_PDF = ['xlsx']
        return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS_PDF

