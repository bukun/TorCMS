import sys
from pathlib import Path
from openpyxl import load_workbook
from xml.sax.saxutils import escape

inws = Path('./xx_outws')

outws = Path('./xx_xml')

if outws.exists():
    pass
else:
    outws.mkdir()

tmpl = open('./tmpl.xml').read()

# print(tmpl)

tmpl_0 = '''<?xml version="1.0" encoding="UTF-8"?>
<csw:Record
	xmlns:csw="http://www.opengis.net/cat/csw/2.0.2"
	xmlns:dc="http://purl.org/dc/elements/1.1/"
	xmlns:dct="http://purl.org/dc/terms/"
	xmlns:ows="http://www.opengis.net/ows"
	xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance"
	xsi:schemaLocation="http://www.opengis.net/cat/csw/2.0.2/record.xsd">
'''

tmpl_9 = '''
</csw:Record>
'''

tp_creator = '<dc:creator>{}</dc:creator>'
tp_contributor = '<dc:contributor>{}</dc:contributor>'
tp_publisher = '<dc:publisher>{}</dc:publisher>'
tp_subject = '<dc:subject>{}</dc:subject>'
tp_abstract = '<dct:abstract>{}</dct:abstract>'
tp_identifier = '<dc:identifier>{}</dc:identifier>'
tp_relation = '<dc:relation>{}</dc:relation>'
tp_source = '<dc:source>{}</dc:source>'
tp_rights = '<dc:rights>{}</dc:rights>'
tp_type = '<dc:type>{}</dc:type>'
tp_title = '<dc:title>{}</dc:title>'
tp_modified = '<dct:modified>{}</dct:modified>'
tp_language = '<dc:language>{}</dc:language>'


# sigs = ['a', 'b', 'c','d','e','f','g','h', 'i', 'j', 'k', 'l', 'm', 'n', 'o','p','q']


def chuli_cell(cell):
    cv = cell.value if cell.value else ''
    return escape(cv)


for xlsx_file in inws.rglob('*.xlsx'):
    if xlsx_file.name.startswith('~$'):
        continue
    try:
        print(xlsx_file.name)
        wb = load_workbook(xlsx_file)
        del (wb)
    except:
        print('Error:', xlsx_file.name)
        sys.exit()

idx = 1
for xlsx_file in inws.rglob('*.xlsx'):

    sig = f'sig{idx}_'

    idx = idx + 1

    print(xlsx_file.name)

    wb = load_workbook(xlsx_file)

    ws = wb.active

    v1 = chuli_cell(ws['B2'])
    v2 = chuli_cell(ws['B41'])
    v3 = chuli_cell(ws['B45'])
    v11 = chuli_cell(ws['B46'])
    v4 = chuli_cell(ws['B39'])
    v5 = chuli_cell(ws['B82'])
    v6 = chuli_cell(ws['B85'])
    v7 = chuli_cell(ws['B44'])
    v8 = chuli_cell(ws['B35'])
    v9 = chuli_cell(ws['B80'])
    v10 = chuli_cell(ws['B35'])

    tt = outws
    if tt.exists():
        pass

    else:
        tt.mkdir()

    outfile = outws / (sig + v1 + '.xml')
    with open(outfile, 'w') as fo:

        fo.write(tmpl_0)
        fo.write('\n')
        fo.write(tp_identifier.format(sig + v1.strip()))
        fo.write('\n')
        fo.write(tp_language.format(v2))
        fo.write('\n')
        fo.write(tp_subject.format(v3 + ',' + v11))
        fo.write('\n')
        fo.write(tp_title.format(v4))
        fo.write('\n')
        fo.write(tp_source.format(v5))
        fo.write('\n')
        fo.write(tp_relation.format(v6))
        fo.write('\n')
        fo.write(tp_abstract.format(v7))
        fo.write('\n')
        fo.write(tp_type.format(v9))
        fo.write('\n')
        fo.write(tp_creator.format(v10))
        fo.write('\n')
        fo.write(tp_contributor.format(v8))

        fo.write(tmpl_9)
