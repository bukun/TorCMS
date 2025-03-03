# 导出仓储数据到excel
import json
import os

import requests
from openpyxl import Workbook

save_file = './xx_outws'

headers = {
    'Content-Type': 'application/json',
    'charset': 'utf-8',
    'User-Agent': 'apifox/1.0.0 (https://www.apifox.cn)',
    'Access-Control-Allow-Origin': '*',  # 这个地方可以写域名也可以是*
    'Access-Control-Allow-Headers': 'x-requested-with',
    'Access-Control-Allow-Methods': 'POST, GET, OPTIONS',
}


def query_data():
    # http://47.243.238.139:2061或http://repository.deep-time.org:2061

    search_payload = {
        'data': 'data',
    }
    search_url = 'http://47.243.238.139:2061/api/ext/v1.0/query/'
    search_d = json.dumps(search_payload)
    search_req_s = requests.post(search_url, data=search_d, headers=headers)
    search_res = json.loads(search_req_s.text)
    id_list = []

    search_payload1 = {'data': 'data', 'pageNum': 1, 'pageSize': search_res['total']}

    search_d1 = json.dumps(search_payload1)
    search_req_s1 = requests.post(search_url, data=search_d1, headers=headers)
    search_res1 = json.loads(search_req_s1.text)

    for info in search_res1['rows']:
        id_list.append(info['datasetId'])

    print("共有记录:" + str(len(id_list)) + "条")
    export_data(id_list)


def export_data(id_list):
    for id in id_list:
        payload = {'datasetId': id}
        search_d = json.dumps(payload)
        url = 'http://47.243.238.139:2061/api/ext/v1.0/query/detail'

        req_s = requests.post(url, data=search_d, headers=headers)
        res = json.loads(req_s.text)
        out_docx = os.path.join(save_file, id + '.xlsx')

        if os.path.exists(save_file):
            pass
        else:
            os.makedirs(save_file)

        if os.path.exists(out_docx):
            try:
                os.remove(out_docx)
            except Exception:
                pass

        wb2 = Workbook()
        ws2 = wb2['Sheet']
        ws2.cell(row=1, column=1).value = 'field_name'
        ws2.cell(row=1, column=2).value = 'value'

        row_inx = 2
        col_inx = 1

        for key, value in res['data'].items():
            ws2.cell(row=row_inx, column=col_inx).value = str(key)
            ws2.cell(row=row_inx, column=col_inx + 1).value = str(value)
            row_inx = row_inx + 1

        wb2.save(out_docx)
        print("*" * 50)
        print("导出仓储数据成功： " + save_file + id + '.xlsx')


if __name__ == '__main__':
    query_data()
