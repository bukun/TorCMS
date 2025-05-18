# -*- coding: utf-8

'''
导入数据集的信息

'''
import os
import pathlib

import tornado.escape
from openpyxl import load_workbook

from torcms.model.category_model import MCategory
from torcms.model.label_model import MPost2Label
from torcms.model.post2catalog_model import MPost2Catalog
from torcms.model.post_model import MPost


def update_category(uid, postdata, kwargs):
    '''
    Update the category of the post.
    '''
    catid = (
        kwargs['catid']
        if ('catid' in kwargs and MCategory.get_by_uid(kwargs['catid']))
        else None
    )

    post_data = postdata

    current_infos = MPost2Catalog.query_by_entity_uid(uid, kind='').objects()

    new_category_arr = []
    # Used to update post2category, to keep order.
    def_cate_arr = ['gcat{0}'.format(x) for x in range(10)]

    # for old page.
    def_cate_arr.append('def_cat_uid')

    # Used to update post extinfo.
    cat_dic = {}
    for key in def_cate_arr:
        if key not in post_data:
            continue
        if post_data[key] == '' or post_data[key] == '0':
            continue
        # 有可能选重复了。保留前面的
        if post_data[key] in new_category_arr:
            continue

        new_category_arr.append(post_data[key] + ' ' * (4 - len(post_data[key])))
        cat_dic[key] = post_data[key] + ' ' * (4 - len(post_data[key]))

    if catid:
        def_cat_id = catid
    elif new_category_arr:
        def_cat_id = new_category_arr[0]
    else:
        def_cat_id = None

    if def_cat_id:
        cat_dic['def_cat_uid'] = def_cat_id
        cat_dic['def_cat_pid'] = MCategory.get_by_uid(def_cat_id).pid

    MPost.update_jsonb(uid, cat_dic)

    for index, catid in enumerate(new_category_arr):
        MPost2Catalog.add_record(uid, catid, index)

    # Delete the old category if not in post requests.
    for cur_info in current_infos:
        if cur_info.tag_id not in new_category_arr:
            MPost2Catalog.remove_relation(uid, cur_info.tag_id)


def chuli_meta(metafile):
    wb = load_workbook(str(metafile))
    sheet = wb[wb.sheetnames[0]]
    meta_dic = {}
    for row in sheet.iter_rows():
        the_key = row[0].value
        the_val = row[1].value if row[1].value else ''
        meta_dic[the_key] = the_val

    cnt_md = meta_dic.get('anytext')

    if cnt_md:
        out_str = ''
        cnts = cnt_md.splitlines()
        for cnt in cnts:
            out_str = out_str + cnt.strip() + '\n\n'

        meta_dic['anytext'] = out_str

    return meta_dic


def get_meta(catid, sig, kind_sig=''):
    '''
    catid : 类别 ID
    sig : Excel 中字段
    kind_sig :
    '''
    if kind_sig:
        pass
    else:
        return
    meta_base = './xx_20221207'
    if os.path.exists(meta_base):
        pass
    else:
        return False

    for wroot, wdirs, wfiles in os.walk(meta_base):
        for wdir in wdirs:
            if wdir.lower().endswith(sig):
                ds_base = pathlib.Path(os.path.join(wroot, wdir))
                return update_db_info(catid, sig, kind_sig, ds_base)


def update_db_info(catid, sig, kind_sig, ds_base):
    pp_data = {'logo': '', 'kind': kind_sig}
    # 给数据新的id。
    sigs = 'd' + sig[-4:]
    kwargsa = {}
    # 对文件夹内容进行遍历，
    #  对不同文件进行处理。
    for uu in ds_base.iterdir():
        if uu.name.endswith('.xlsx'):
            meta_dic = chuli_meta(uu)

            # print(meta_dic.keys())
            if 'title' in meta_dic:
                pass
            else:
                continue
            pp_data['title'] = meta_dic['title']
            # pp_data['cnt_md'] = meta_dic['anytext']
            pp_data['user_name'] = 'admin'
            pp_data['user_name'] = 'admin'
            pp_data['cnt_md'] = meta_dic['anytext']
            pp_data['tags'] = meta_dic['keywords']
            pp_data['gcat0'] = catid
            pp_data['def_cat_pid'] = catid[:2] + '00'
            # 将主要数据添加到外扩展
            pp_data['extinfo'] = {}
            for key in meta_dic.keys():
                if key != 'field_name':
                    pp_data['extinfo']['pycsw_' + str(key)] = meta_dic.get(key, '')

            kwargsa = {
                'gcat0': catid,
                'cat_id': catid,
            }
    print(pp_data)
    post_id = sigs
    postinfo = MPost.get_by_uid(post_id)
    if postinfo:
        postinfo.extinfo.update(pp_data['extinfo'])
        pp_data['extinfo'] = postinfo.extinfo
        # 说明内容可能会添加补充其他信息。
        pp_data['cnt_md'] = tornado.escape.xhtml_unescape(postinfo.cnt_md)

    else:
        MPost.add_or_update(sigs, pp_data)

    MPost.add_or_update(sigs, pp_data, update_time=False)
    update_category(sigs, pp_data, kwargsa)
    update_label(sigs, pp_data)


def update_label(signature, post_data):
    '''
    Update the label .
    '''
    current_tag_infos = MPost2Label.get_by_uid(signature).objects()
    if 'tags' in post_data:
        pass
    else:
        return False

    if '；' in post_data['tags']:
        tags_arr = [x.strip() for x in post_data['tags'].split('；')]
    elif ',' in post_data['tags']:
        tags_arr = [x.strip() for x in post_data['tags'].split(',')]
    elif '，' in post_data['tags']:
        tags_arr = [x.strip() for x in post_data['tags'].split('，')]
    elif ';' in post_data['tags']:
        tags_arr = [x.strip() for x in post_data['tags'].split(';')]
    else:
        tags_arr = [x.strip() for x in post_data['tags'].split(' ')]

    if len(tags_arr) > 5:
        del tags_arr[5:]

    for tag_name in tags_arr:
        if tag_name == '':
            pass
        else:
            MPost2Label.add_record(signature, tag_name, 1)

    for cur_info in current_tag_infos:
        if cur_info.tag_name in tags_arr:
            pass
        else:
            MPost2Label.remove_relation(signature, cur_info.tag_id)


def import_meta():
    # 此文件夹下声明系统中的数据集及分类

    dts_path_obj = pathlib.Path('database/datasets')
    for wfile in dts_path_obj.rglob('*.xlsx'):
        chli_xlsx(wfile)


def chli_xlsx(cat_path):
    '''
    处理网站中的数据集列表的 XLSX 文件
    '''
    full_path_str = str(cat_path.resolve())

    wb = load_workbook(full_path_str)

    sheets = wb.sheetnames
    for sheet_name in sheets:
        catid = sheet_name.split('-')[0]
        ws = wb[sheet_name]
        # rows = ws.max_row
        # cols = ws.max_column
        for row in ws.rows:
            # print(type(row))
            sig = row[0].value
            # print(sig)
            if sig:
                get_meta(catid, sig, kind_sig='d')


if __name__ == '__main__':
    import_meta()
