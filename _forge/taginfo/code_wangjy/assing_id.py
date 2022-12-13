from openpyxl import load_workbook
inw = 'meta20210222.xlsx'
out_xlsx = 'meta20211101.xlsx'
wb_new = load_workbook(out_xlsx)
ws_new = wb_new.active

wb = load_workbook(inw)
ws = wb.active
row_idx = 2
for row in ws.iter_rows(min_row=2):
    meta_dic = {}
    meta_dic['id'] = row[0].value
    meta_dic['cpbh'] = row[1].value  # 产品编号
    meta_dic['cpmc'] = row[2].value  # 产品名称
    meta_dic['cplx'] = row[3].value  # 产品类型
    meta_dic['cp_abs'] = row[4].value  # 简介
    meta_dic['xkfl'] = row[5].value  # 学科分类

    for row_new in ws_new.iter_rows(min_row = 2):
        if row_new[1].value == meta_dic['cpbh']:
            row_new[0].value = meta_dic['id']
            break

    row_idx  = row_idx + 1
    print(meta_dic)

wb_new.save('xx_out.xlsx')