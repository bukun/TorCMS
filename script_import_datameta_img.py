# -*- coding: utf-8

'''
导入数据集IMG
'''
import os

import pathlib
from openpyxl import load_workbook
from torcms.model.category_model import TabPost
from PIL import Image

logo_cache_dir = './static/cache'

def update_logo(logo,uid):
    entry = TabPost.update(
        logo = logo
    ).where(TabPost.uid == uid)
    entry.execute()
    return True


def get_meta(catid, sig, kind_sig=''):
    '''
    catid : 类别 ID
    sig : Excel 中字段
    kind_sig :
    '''
    if kind_sig:
        pass
    else:
        return
    meta_base = './xx_20221207'
    if os.path.exists(meta_base):
        pass
    else:
        return False

    for wroot, wdirs, wfiles in os.walk(meta_base):
        for wdir in wdirs:
            if wdir.lower().endswith(sig):
                ds_base = pathlib.Path(os.path.join(wroot, wdir))
                return update_db_info(catid, sig, kind_sig, ds_base)


def update_db_info(catid, sig, kind_sig, ds_base):
    pp_data = {'logo': '', 'kind': kind_sig}
    # 给数据新的id。
    sigs = 'd' + sig[-4:]
    kwargsa = { }
    # 对文件夹内容进行遍历，
    #  对不同文件进行处理。
    for uu in ds_base.iterdir():
        if uu.name.startswith('thumbnail_'):
            hou = os.path.splitext(uu.name)[-1]
            if hou in ['.jpg', '.png', '.JPG', '.PNG', 'jpeg', 'JPEG']:
                thum_path = gen_thumb(ds_base / uu.name,sig[-4:])
                pp_data['logo'] = thum_path.strip('.')

    print(pp_data)
    update_logo(pp_data['logo'],sigs)


def gen_thumb(img_path, sig):
    img = Image.open(img_path)
    img_width = 400 if img.size[0] > 400 else img.size[0]
    img_height = 300 if img.size[1] > 300 else img.size[1]
    img.thumbnail((img_width, img_height), Image.ANTIALIAS)
    if os.path.exists(logo_cache_dir):
        pass
    else:
        os.mkdir(logo_cache_dir)
    try:
        thum_file_path = os.path.join(logo_cache_dir, 'd' + sig + '.jpg')
        img.save(thum_file_path, "JPEG")
    except Exception:
        thum_file_path = os.path.join(logo_cache_dir, 'd' + sig + '.png')
        img.save(thum_file_path, "PNG")
    return thum_file_path

def import_meta():
    # 此文件夹下声明系统中的数据集及分类

    dts_path_obj = pathlib.Path('./database/datasets')
    for wfile in dts_path_obj.rglob('*.xlsx'):
        chli_xlsx(wfile)


def chli_xlsx(cat_path):
    full_path_str = str(cat_path.resolve())

    wb = load_workbook(full_path_str)

    sheets = wb.sheetnames
    for sheet_name in sheets:
        catid = sheet_name.split('-')[0]
        ws = wb[sheet_name]
        for row in ws.rows:
            sig = row[0].value
            if sig:
                get_meta(catid, sig, kind_sig='d')



if __name__ == '__main__':
    import_meta()
